"""Exporta los candidatos de matching a Excel para validación humana.

Hojas:
  1. Resumen                → conteos, advertencias, distribución
  2. Matches_1a1            → los 8 matches de alta confianza (para verificar 1x1)
  3. Ambiguos               → candidatos con ties (necesitan decisión humana)
  4. Top_500_alta_proba     → 500 candidatos ordenados por proba (panel de revisión)
  5. Por_cluster            → agrupados por cluster bi-tienda

Columna "decision" con validación (Sí / No / Duda) para que el usuario marque.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data")
IN_CANDIDATOS = BASE / "matches_candidatos.json"
IN_TOP        = BASE / "matches_top.json"
OUT_XLSX      = BASE / "matches_revisar.xlsx"

# Colores
COLOR_HEADER    = "37474F"
COLOR_MATCH     = "C8E6C9"   # verde suave para matches 1-a-1
COLOR_AMBIGUO   = "FFE0B2"   # naranja suave para ambiguos
COLOR_ALTA      = "E8F5E9"   # verde muy suave
COLOR_MEDIA     = "FFF9C4"   # amarillo
COLOR_BAJA      = "FFEBEE"   # rojo suave
COLOR_ZEBRA     = "F5F5F5"

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITLE   = Font(name="Arial", size=14, bold=True)

THIN = Side(border_style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    "decisión",
    "proba", "cosine", "cluster",
    "easy_marca", "easy_título", "easy_precio",
    "sodi_marca", "sodi_título", "sodi_precio",
    "easy_sku", "sodi_sku",
]
WIDTHS = [11, 8, 8, 9, 14, 55, 12, 14, 55, 12, 14, 14]


def _color_proba(p: float) -> str:
    if p >= 0.90: return COLOR_ALTA
    if p >= 0.70: return COLOR_MEDIA
    return COLOR_BAJA


def _header(ws, row, headers):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fila(rec, color=None):
    return [
        "",
        rec["proba"], rec["cosine"], rec["cluster_id"],
        rec.get("easy_marca") or "",
        rec.get("easy_titulo") or "",
        rec.get("easy_precio"),
        rec.get("sodimac_marca") or "",
        rec.get("sodimac_titulo") or "",
        rec.get("sodimac_precio"),
        rec.get("easy_sku") or "",
        rec.get("sodimac_sku") or "",
    ]


def _escribir_data(ws, rows, start_row=2, color_fn=None):
    for i, rec in enumerate(rows, start=start_row):
        fila = _fila(rec)
        color = color_fn(rec) if color_fn else _color_proba(rec["proba"])
        fill = PatternFill("solid", fgColor=color)
        for col, v in enumerate(fila, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_DEFAULT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            c.fill = fill
    # validación en col A (decisión)
    n = len(rows)
    if n:
        dv = DataValidation(type="list", formula1='"✓ match,✗ no match,? duda"', allow_blank=True)
        dv.add(f"A{start_row}:A{start_row+n-1}")
        ws.add_data_validation(dv)


def _hoja_resumen(wb, candidatos, top_data):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20

    r = 1
    ws.cell(row=r, column=1, value="MATCHES EASY↔SODIMAC - PANEL DE REVISIÓN").font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    meta = top_data["meta"]
    filas = [
        ("Total candidatos (proba ≥ 0.50)",    meta["total_candidatos"]),
        ("Matches 1-a-1 (alta confianza ≥ 0.85)", meta["matches_1a1"]),
        ("Ambiguos (ties)",                    meta["ambiguos"]),
        ("Easy matcheados",                    f"{meta['easy_matcheados']} / 2221"),
        ("Sodimac matcheados",                 f"{meta['sodi_matcheados']} / 1421"),
        ("Umbral alta confianza",              meta["umbral_alta_conf"]),
        ("Umbral candidato",                   meta["umbral_candidato"]),
    ]
    for k, v in filas:
        ws.cell(row=r, column=1, value=k).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=2, value=v).font = FONT_DEFAULT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="⚠  ADVERTENCIA: DATA LEAKAGE").font = Font(name="Arial", size=12, bold=True, color="C62828")
    r += 1
    mensaje = (
        "El modelo RF alcanzó F1=1.0 en validación cruzada porque usa los mismos "
        "features que las reglas de weak-labeling (cosine_sim, brand_match, "
        "jaccard_titulo). Esto NO garantiza que generalice. Es imprescindible "
        "validar manualmente los matches en este Excel antes de usarlos."
    )
    c = ws.cell(row=r, column=1, value=mensaje)
    c.font = FONT_DEFAULT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=2)
    ws.row_dimensions[r].height = 18
    r += 6

    ws.cell(row=r, column=1, value="CÓMO USAR ESTE EXCEL").font = Font(name="Arial", size=12, bold=True)
    r += 1
    instr = [
        "1. En cada hoja, revisa los matches propuestos Easy↔Sodimac.",
        "2. En la columna 'decisión' elige: ✓ match | ✗ no match | ? duda",
        "3. Ordena por proba descendente para priorizar los más seguros.",
        "4. Con las decisiones humanas podemos re-entrenar el modelo.",
    ]
    for t in instr:
        ws.cell(row=r, column=1, value=t).font = FONT_DEFAULT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DISTRIBUCIÓN DE PROBABILIDADES").font = Font(name="Arial", size=12, bold=True)
    r += 1
    _header(ws, r, ["rango_proba", "n_candidatos"])
    ws.column_dimensions["B"].width = 20
    r += 1
    rangos = [(0.95, 1.01), (0.85, 0.95), (0.70, 0.85), (0.50, 0.70)]
    for lo, hi in rangos:
        n = sum(1 for c in candidatos if lo <= c["proba"] < hi)
        ws.cell(row=r, column=1, value=f"{lo:.2f} ≤ p < {hi:.2f}").font = FONT_DEFAULT
        ws.cell(row=r, column=2, value=n).font = FONT_DEFAULT
        r += 1


def _hoja_matches_1a1(wb, matches):
    ws = wb.create_sheet("Matches_1a1")
    _widths(ws, WIDTHS)
    _header(ws, 1, COLS)
    ws.freeze_panes = "A2"
    _escribir_data(ws, matches, color_fn=lambda r: COLOR_MATCH)


def _hoja_ambiguos(wb, ambiguos):
    ws = wb.create_sheet("Ambiguos")
    _widths(ws, WIDTHS)
    _header(ws, 1, COLS)
    ws.freeze_panes = "A2"
    ambiguos.sort(key=lambda r: -r["proba"])
    _escribir_data(ws, ambiguos[:500], color_fn=lambda r: COLOR_AMBIGUO)


def _hoja_top_500(wb, candidatos):
    ws = wb.create_sheet("Top_500_alta_proba")
    _widths(ws, WIDTHS)
    _header(ws, 1, COLS)
    ws.freeze_panes = "A2"
    candidatos.sort(key=lambda r: -r["proba"])
    _escribir_data(ws, candidatos[:500])


def _hoja_por_cluster(wb, candidatos):
    """Agrupa los top matches por cluster bi-tienda."""
    from collections import defaultdict
    por_cluster = defaultdict(list)
    for c in candidatos:
        por_cluster[c["cluster_id"]].append(c)

    ws = wb.create_sheet("Por_cluster")
    _widths(ws, WIDTHS)
    _header(ws, 1, COLS)
    ws.freeze_panes = "A2"

    r = 2
    # clusters ordenados por número de candidatos
    for cid in sorted(por_cluster.keys(), key=lambda k: -len(por_cluster[k])):
        filas = por_cluster[cid]
        if len(filas) < 2:
            continue
        filas.sort(key=lambda x: -x["proba"])
        # separador de cluster
        fill_sep = PatternFill("solid", fgColor="B3E5FC")
        c = ws.cell(row=r, column=1, value=f"═══ CLUSTER {cid}  ({len(filas)} candidatos, proba máx={filas[0]['proba']:.3f})")
        c.font = Font(name="Arial", size=11, bold=True)
        c.fill = fill_sep
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        r += 1
        _escribir_data(ws, filas[:30], start_row=r)
        r += len(filas[:30]) + 1


def main():
    print("=" * 70)
    print("EXPORTANDO MATCHES A EXCEL (PANEL DE REVISIÓN)")
    print("=" * 70)

    with open(IN_CANDIDATOS, encoding="utf-8") as f:
        candidatos = json.load(f)
    with open(IN_TOP, encoding="utf-8") as f:
        top_data = json.load(f)

    print(f"  Candidatos cargados: {len(candidatos)}")
    print(f"  Matches 1-a-1     : {len(top_data['matches_1a1'])}")
    print(f"  Ambiguos          : {len(top_data['ambiguos'])}")

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    _hoja_resumen(wb, candidatos, top_data)
    _hoja_matches_1a1(wb, top_data["matches_1a1"])
    _hoja_ambiguos(wb, top_data["ambiguos"])
    _hoja_top_500(wb, candidatos)
    _hoja_por_cluster(wb, candidatos)

    wb.save(OUT_XLSX)
    size_kb = OUT_XLSX.stat().st_size // 1024
    print(f"\n  Excel guardado: {OUT_XLSX}")
    print(f"  Tamaño: {size_kb} KB")
    print(f"  Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
