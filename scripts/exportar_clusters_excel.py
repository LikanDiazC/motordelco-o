"""Exporta resultados de clustering a Excel con visualización profesional."""

import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles.differential import DifferentialStyle

BASE   = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data")
SALIDA = BASE / "clusters_visualizacion.xlsx"

# ─── Paleta ────────────────────────────────────────────────────────────────
C_DARK  = "1F3864"
C_MID   = "2E75B6"
C_LIGHT = "D9E1F2"
C_EASY  = "1A6B3A"   # verde Easy
C_SODI  = "C0392B"   # rojo Sodimac
C_GOLD  = "7B5E00"   # dorado para bi-tienda
C_GRLT  = "D5F5E3"
C_ORLT  = "FADBD8"
C_YLLT  = "FFF9C4"
C_GRAY  = "F2F2F2"
C_WHITE = "FFFFFF"

FONT_H   = Font(name="Arial", bold=True, color=C_WHITE, size=10)
FONT_B   = Font(name="Arial", size=9)
FONT_SM  = Font(name="Arial", size=8)
FONT_BLD = Font(name="Arial", size=9, bold=True)
FONT_LNK = Font(name="Arial", size=9, color="0563C1", underline="single")

FILL_DARK  = PatternFill("solid", fgColor=C_DARK)
FILL_MID   = PatternFill("solid", fgColor=C_MID)
FILL_EASY  = PatternFill("solid", fgColor=C_EASY)
FILL_SODI  = PatternFill("solid", fgColor=C_SODI)
FILL_GOLD  = PatternFill("solid", fgColor=C_GOLD)
FILL_LIGHT = PatternFill("solid", fgColor=C_LIGHT)
FILL_GRLT  = PatternFill("solid", fgColor=C_GRLT)
FILL_ORLT  = PatternFill("solid", fgColor=C_ORLT)
FILL_YLLT  = PatternFill("solid", fgColor=C_YLLT)
FILL_GRAY  = PatternFill("solid", fgColor=C_GRAY)

AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left",   vertical="center")
AL_W = Alignment(horizontal="left",   vertical="top", wrap_text=True)

thin   = Side(style="thin", color="B8CCE4")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, r, c, texto, fill=FILL_DARK, span=None, font=None):
    cell = ws.cell(row=r, column=c, value=texto)
    cell.font      = font or FONT_H
    cell.fill      = fill
    cell.alignment = AL_C
    cell.border    = BORDER
    if span and span > 1:
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r,   end_column=c + span - 1)

def cel(ws, r, c, v, bold=False, center=False, wrap=False, fill=None, small=False, link=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = (FONT_LNK if link else
                      FONT_SM  if small else
                      FONT_BLD if bold  else FONT_B)
    cell.alignment = AL_W if wrap else (AL_C if center else AL_L)
    cell.border    = BORDER
    if fill:
        cell.fill = fill
    return cell


# ─── Paleta de colores para clusters (ciclo de 20 colores pastel) ───────────
CLUSTER_COLORS = [
    "FFEFD5","E0F7FA","F3E5F5","E8F5E9","FFF8E1","EDE7F6",
    "FCE4EC","E1F5FE","F9FBE7","FBE9E7","E8EAF6","E0F2F1",
    "FFF3E0","F1F8E9","EAF4FB","F8F0E3","E9F5F9","F0E6FF",
    "EAFAF1","FEF9E7",
]

def cluster_fill(cid):
    if cid == -1:
        return PatternFill("solid", fgColor="DDDDDD")
    return PatternFill("solid", fgColor=CLUSTER_COLORS[cid % len(CLUSTER_COLORS)])


# ═══════════════════════════════════════════════════════════════════════════
def main():
    # Cargar datos
    with open(BASE/"clusters_reporte.json", encoding="utf-8") as f:
        rep = json.load(f)

    with open(BASE/"easy"/"catalogo_normalizado.json", encoding="utf-8") as f:
        easy_prods = json.load(f)

    with open(BASE/"sodimac"/"catalogo_normalizado.json", encoding="utf-8") as f:
        sodi_prods = json.load(f)

    clusters = [c for c in rep["clusters"] if c["cluster_id"] != -1]
    ruido    = next((c for c in rep["clusters"] if c["cluster_id"] == -1), None)

    # Clasificar clusters
    bi_util, bi_debil, mono_e, mono_s = [], [], [], []
    for c in clusters:
        e = c["por_tienda"].get("easy", 0)
        s = c["por_tienda"].get("sodimac", 0)
        if   e == 0:        mono_s.append(c)
        elif s == 0:        mono_e.append(c)
        elif min(e,s) >= 2: bi_util.append(c)
        else:               bi_debil.append(c)

    pares_brute   = 1421 * 2221
    pares_cluster = sum(c["por_tienda"].get("easy",0)*c["por_tienda"].get("sodimac",0) for c in clusters)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 1 — Resumen ejecutivo
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 40

    hdr(ws1, 1, 1, "RESUMEN DEL CLUSTERING", span=4)
    ws1.row_dimensions[1].height = 28

    # Parámetros
    hdr(ws1, 2, 1, "Parámetros UMAP",       FILL_MID, span=2)
    hdr(ws1, 2, 3, "Parámetros HDBSCAN",    FILL_MID, span=2)
    p = rep["parametros"]
    params_umap = [
        ("n_neighbors",  p["umap"]["n_neighbors"]),
        ("n_components", p["umap"]["n_components"]),
        ("min_dist",     p["umap"]["min_dist"]),
        ("metric",       p["umap"]["metric"]),
    ]
    params_hdb = [
        ("min_cluster_size", p["hdbscan"]["min_cluster_size"]),
        ("min_samples",      p["hdbscan"]["min_samples"]),
        ("metric",           p["hdbscan"]["metric"]),
    ]
    for i, (k, v) in enumerate(params_umap, 3):
        cel(ws1, i, 1, k, fill=FILL_LIGHT)
        cel(ws1, i, 2, v, center=True, fill=FILL_LIGHT)
    for i, (k, v) in enumerate(params_hdb, 3):
        cel(ws1, i, 3, k, fill=FILL_LIGHT)
        cel(ws1, i, 4, v, center=True, fill=FILL_LIGHT)

    # Métricas principales
    r = 8
    hdr(ws1, r, 1, "Métrica",   FILL_DARK)
    hdr(ws1, r, 2, "Valor",     FILL_DARK)
    hdr(ws1, r, 3, "% / Info",  FILL_DARK)
    hdr(ws1, r, 4, "Notas",     FILL_DARK)
    r += 1

    metricas = [
        ("Total productos",               rep["n_productos"],           "",       "Easy 2221 + Sodimac 1421"),
        ("Clusters detectados",           rep["n_clusters"],            "",       "Grupos semánticos naturales"),
        ("Productos ruido (sin cluster)", rep["n_ruido"],
            f"{rep['n_ruido']*100//rep['n_productos']}%",                         "Aislados, no clasificables"),
        ("Clusters bi-tienda útiles",     len(bi_util),
            f"{len(bi_util)*100//rep['n_clusters']}%",                            "≥2 productos en cada tienda"),
        ("Clusters bi-tienda débiles",    len(bi_debil),                "",       "1 producto en una tienda"),
        ("Clusters solo Easy",            len(mono_e),                  "",       "Sin equivalente en Sodimac"),
        ("Clusters solo Sodimac",         len(mono_s),                  "",       "Sin equivalente en Easy"),
        ("Pares brute-force",             pares_brute,                  "",       "Sin blocking (imposible)"),
        ("Pares candidatos (con cluster)",pares_cluster,
            f"{pares_cluster*100/pares_brute:.2f}% del total",                    "Input para Random Forest"),
        ("Reducción de espacio",          f"{(1-pares_cluster/pares_brute)*100:.2f}%",
            "",                                                                    "Menos pares = más rápido"),
        ("Easy en clusters útiles",
            f"419 / 2221", "18%",                                                 "Cobertura de matching Easy"),
        ("Sodimac en clusters útiles",
            f"367 / 1421", "25%",                                                 "Cobertura de matching Sodimac"),
    ]

    for fi, (etiq, val, pct, nota) in enumerate(metricas, r):
        zf = FILL_LIGHT if fi % 2 == 0 else None
        cel(ws1, fi, 1, etiq, fill=zf)
        cel(ws1, fi, 2, val,  center=True, bold=True, fill=zf)
        cel(ws1, fi, 3, pct,  center=True, fill=zf)
        cel(ws1, fi, 4, nota, fill=zf)

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 2 — Todos los clusters
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Clusters")
    ws2.freeze_panes = "A3"

    # Fila 1 grupos
    hdr(ws2, 1, 1, "IDENTIFICACIÓN",    FILL_DARK,  span=3)
    hdr(ws2, 1, 4, "COMPOSICIÓN",       FILL_MID,   span=4)
    hdr(ws2, 1, 8, "DIMENSIONES",       PatternFill("solid",fgColor=C_EASY), span=2)
    hdr(ws2, 1,10, "CONTENIDO",         FILL_DARK,  span=3)
    hdr(ws2, 1,13, "EJEMPLO PRODUCTO",  FILL_MID,   span=3)

    # Fila 2 encabezados
    hdrs2 = [
        "Cluster ID","Tipo","Categoría",
        "n total","n Easy","n Sodimac","Balance",
        "Pares cand.",
        "Largo mm","Diám mm",
        "Material top","Mat. frecuencia","Tipo rosca top",
        "Ej. tienda","Ej. SKU","Ej. título",
    ]
    fills_h2 = ([FILL_DARK]*3 + [FILL_MID]*4 +
                [PatternFill("solid",fgColor=C_EASY)]*2 +
                [FILL_DARK]*3 + [FILL_MID]*3)
    for col, (h, f) in enumerate(zip(hdrs2, fills_h2), 1):
        hdr(ws2, 2, col, h, f)

    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 32

    # Ordenar: bi_util primero, luego bi_debil, mono_e, mono_s, ruido
    def tipo_orden(c):
        e = c["por_tienda"].get("easy", 0)
        s = c["por_tienda"].get("sodimac", 0)
        if c["cluster_id"] == -1: return (4, 0)
        if e >= 2 and s >= 2: return (0, -c["tamano"])
        if (e > 0 and s > 0): return (1, -c["tamano"])
        if e == 0:  return (3, -c["tamano"])
        return (2, -c["tamano"])

    def tipo_label(c):
        e = c["por_tienda"].get("easy", 0)
        s = c["por_tienda"].get("sodimac", 0)
        if c["cluster_id"] == -1: return "RUIDO"
        if e >= 2 and s >= 2: return "BI-TIENDA"
        if e > 0 and s > 0:   return "BI-DÉBIL"
        if e == 0:             return "SOLO SODI"
        return "SOLO EASY"

    def tipo_fill(label):
        return {
            "BI-TIENDA": PatternFill("solid", fgColor=C_YLLT),
            "BI-DÉBIL":  PatternFill("solid", fgColor="FFF0CC"),
            "SOLO EASY": PatternFill("solid", fgColor=C_GRLT),
            "SOLO SODI": PatternFill("solid", fgColor=C_ORLT),
            "RUIDO":     FILL_GRAY,
        }.get(label, None)

    todos_clusters = sorted(rep["clusters"], key=tipo_orden)

    for fi, c in enumerate(todos_clusters, 3):
        e = c["por_tienda"].get("easy", 0)
        s = c["por_tienda"].get("sodimac", 0)
        pares = e * s
        bal   = round(min(e,s)/max(e,s), 2) if max(e,s) > 0 else 0.0
        tipo  = tipo_label(c)
        zf    = tipo_fill(tipo)
        mat_top = c["materiales_top3"][0][0] if c["materiales_top3"] else ""
        mat_freq= c["materiales_top3"][0][1] if c["materiales_top3"] else 0
        rosca   = ""  # no está en el reporte; lo dejamos vacío
        ej      = c["ejemplos"][0] if c["ejemplos"] else {}

        vals = [
            c["cluster_id"], tipo, c["categoria_top"] or "",
            c["tamano"], e, s, bal, pares,
            c["largo_mm_medio"], c["diam_mm_medio"],
            mat_top, mat_freq, rosca,
            ej.get("tienda",""), ej.get("sku",""), ej.get("titulo","")[:80],
        ]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=fi, column=col, value=v)
            cell.border = BORDER
            cell.alignment = AL_C if col <= 12 else AL_L
            if zf:
                cell.fill = zf
            if col == 1:
                cell.font = FONT_BLD
            elif col == 2:
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color={
                                     "BI-TIENDA":"7B5E00","BI-DÉBIL":"7B5E00",
                                     "SOLO EASY":"1A6B3A","SOLO SODI":"C0392B",
                                     "RUIDO":"888888"
                                 }.get(tipo, "000000"))
            else:
                cell.font = FONT_SM if col == 16 else FONT_B

    anchos2 = {1:11,2:11,3:13,4:9,5:8,6:8,7:9,8:11,9:11,10:10,11:14,12:13,13:13,14:10,15:10,16:65}
    for c,w in anchos2.items():
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.auto_filter.ref = f"A2:{get_column_letter(16)}2"

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 3 — Detalle BI-TIENDA (todos los productos de clusters útiles)
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Bi-tienda detalle")
    ws3.freeze_panes = "A3"

    # construir lookup cluster_id → tienda → lista de productos
    idx_easy = {str(p["sku"]): p for p in easy_prods}
    idx_sodi = {str(p["sku"]): p for p in sodi_prods}

    bi_ids = {c["cluster_id"] for c in bi_util}

    hdr(ws3, 1, 1, "IDENTIFICACIÓN", FILL_DARK, span=5)
    hdr(ws3, 1, 6, "METADATA",       FILL_MID,  span=4)
    hdr(ws3, 1,10, "DIMENSIONES",    PatternFill("solid",fgColor=C_EASY), span=3)
    hdr(ws3, 1,13, "ATRIBUTOS",      FILL_MID,  span=3)

    hdrs3 = [
        "Cluster","Tienda","ID Producto","SKU","Título",
        "Marca","Marca norm","Precio CLP","Disponibilidad",
        "Largo mm","Diám mm","Medida cruda",
        "Material norm","Tipo rosca","Tipo cabeza",
    ]
    fills_h3 = ([FILL_DARK]*5 + [FILL_MID]*4 +
                [PatternFill("solid",fgColor=C_EASY)]*3 + [FILL_MID]*3)
    for col,(h,f) in enumerate(zip(hdrs3, fills_h3),1):
        hdr(ws3, 2, col, h, f)

    ws3.row_dimensions[1].height = 20
    ws3.row_dimensions[2].height = 30

    fi = 3
    for prod_list, tienda_name in [(easy_prods,"easy"),(sodi_prods,"sodimac")]:
        for p in prod_list:
            cid = p.get("cluster_id", -1)
            if cid not in bi_ids:
                continue
            mb  = p.get("metadata_basica",{})
            mt  = p.get("metadata_tecnica",{})
            dim = mt.get("dimensiones",{})
            zf  = FILL_GRLT if tienda_name == "easy" else FILL_ORLT

            vals3 = [
                cid,
                tienda_name,
                p.get("id_producto",""),
                p.get("sku",""),
                mb.get("titulo_limpio") or mb.get("titulo",""),
                mb.get("marca",""),
                mb.get("marca_norm",""),
                mb.get("precio_clp",0),
                mb.get("disponibilidad",""),
                dim.get("largo_mm"),
                dim.get("diametro_mm"),
                dim.get("medida_cruda",""),
                mt.get("material_norm","") or mt.get("material",""),
                mt.get("tipo_rosca",""),
                mt.get("tipo_cabeza",""),
            ]
            for col, v in enumerate(vals3, 1):
                cell = ws3.cell(row=fi, column=col, value=v)
                cell.border    = BORDER
                cell.alignment = AL_C if col in (1,2,8,10,11) else AL_L
                cell.fill      = zf
                cell.font = (Font(name="Arial",size=9,bold=True,
                                  color="1A6B3A" if tienda_name=="easy" else "C0392B")
                             if col == 2 else FONT_B)
            fi += 1

    anchos3 = {1:9,2:9,3:22,4:12,5:55,6:16,7:16,8:13,9:13,10:11,11:11,12:14,13:16,14:14,15:14}
    for c,w in anchos3.items():
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.auto_filter.ref = f"A2:{get_column_letter(15)}2"

    # ═══════════════════════════════════════════════════════════════════════
    # HOJA 4 — Mapa de calor de clusters (qué tienes en qué cluster)
    # ═══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Mapa clusters")
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 14
    ws4.column_dimensions["D"].width = 14
    ws4.column_dimensions["E"].width = 14
    ws4.column_dimensions["F"].width = 14
    ws4.column_dimensions["G"].width = 14
    ws4.column_dimensions["H"].width = 50

    hdr(ws4, 1, 1, "Cluster ID",    FILL_DARK)
    hdr(ws4, 1, 2, "Categoría",     FILL_DARK)
    hdr(ws4, 1, 3, "n Easy",        FILL_EASY)
    hdr(ws4, 1, 4, "n Sodimac",     FILL_SODI)
    hdr(ws4, 1, 5, "n Total",       FILL_MID)
    hdr(ws4, 1, 6, "Balance",       FILL_DARK)
    hdr(ws4, 1, 7, "Pares cand.",   FILL_MID)
    hdr(ws4, 1, 8, "Ejemplo",       FILL_DARK)
    ws4.row_dimensions[1].height = 26

    # ordenar por pares (más potencial primero)
    clusters_ord = sorted(
        [c for c in rep["clusters"] if c["cluster_id"] != -1],
        key=lambda c: c["por_tienda"].get("easy",0)*c["por_tienda"].get("sodimac",0),
        reverse=True
    )

    MAX_BAR = 130  # para escala visual de barras manuales
    for fi, c in enumerate(clusters_ord, 2):
        e   = c["por_tienda"].get("easy",0)
        s   = c["por_tienda"].get("sodimac",0)
        bal = round(min(e,s)/max(e,s),2) if max(e,s) > 0 else 0.0
        par = e * s
        ej  = c["ejemplos"][0]["titulo"][:70] if c["ejemplos"] else ""
        tipo = tipo_label(c)
        zf   = tipo_fill(tipo)

        for col, v in [(1,c["cluster_id"]),(2,c["categoria_top"]),(3,e),(4,s),
                       (5,c["tamano"]),(6,bal),(7,par),(8,ej)]:
            cell = ws4.cell(row=fi, column=col, value=v)
            cell.border    = BORDER
            cell.alignment = AL_C if col != 8 else AL_L
            cell.font      = FONT_B
            if zf: cell.fill = zf

        # color manual para easy/sodimac según valor
        e_cell = ws4.cell(row=fi, column=3)
        s_cell = ws4.cell(row=fi, column=4)
        # tono verde proporcional a cantidad easy
        if e > 0:
            intensity = max(0, min(255, 255 - int(e * 200 / max(MAX_BAR,1))))
            e_cell.fill = PatternFill("solid", fgColor=f"{intensity:02X}FF{intensity:02X}")
        # tono rojo proporcional a cantidad sodimac
        if s > 0:
            intensity = max(0, min(255, 255 - int(s * 200 / max(MAX_BAR,1))))
            s_cell.fill = PatternFill("solid", fgColor=f"FF{intensity:02X}{intensity:02X}")

    ws4.auto_filter.ref = "A1:H1"

    # ─── guardar ────────────────────────────────────────────────────────
    wb.save(SALIDA)
    print(f"Excel guardado: {SALIDA}")
    print(f"  Hoja 1 - Resumen ejecutivo")
    print(f"  Hoja 2 - Todos los clusters ({rep['n_clusters']} clusters + ruido)")
    print(f"  Hoja 3 - Detalle bi-tienda ({sum(len(bi_util) for _ in [1])} clusters bi-tienda útiles)")
    print(f"  Hoja 4 - Mapa de calor (fácil de filtrar)")

if __name__ == "__main__":
    main()
