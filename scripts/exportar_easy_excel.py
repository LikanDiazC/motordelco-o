"""Exporta catalogo_profundo.json de Easy a Excel con formato profesional."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA  = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data\easy")
SALIDA = DATA / "catalogo_easy.xlsx"

SPECS_COLUMNAS = [
    "Tipo de producto",
    "Modelo",
    "Contenido",
    "Tono",
    "Terminación",
    "Materiales",
    "Material",
    "Diámetro",
    "Largo",
    "Uso",
    "Uso Recomendado",
    "Origen",
    "País de Origen",
    "Garantía Mínima Legal",
    "Observaciones y recomendaciones",
]

# ---------------------------------------------------------------------------
# Estilos  (misma paleta que Sodimac pero en verde Easy)
# ---------------------------------------------------------------------------
COLOR_DARK  = "1A5276"   # verde-azul oscuro (Easy)
COLOR_MID   = "1F618D"   # azul medio
COLOR_LIGHT = "D6EAF8"   # celeste muy claro (zebra)
COLOR_WHITE = "FFFFFF"
COLOR_BORD  = "A9CCE3"

FONT_H = Font(name="Arial", bold=True, color=COLOR_WHITE, size=10)
FONT_B = Font(name="Arial", size=9)
FONT_L = Font(name="Arial", size=9, color="0563C1", underline="single")
FONT_BOLD = Font(name="Arial", size=9, bold=True)

FILL_DARK  = PatternFill("solid", fgColor=COLOR_DARK)
FILL_MID   = PatternFill("solid", fgColor=COLOR_MID)
FILL_LIGHT = PatternFill("solid", fgColor=COLOR_LIGHT)

ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left",   vertical="center")
ALIGN_W = Alignment(horizontal="left",   vertical="top", wrap_text=True)

thin   = Side(style="thin", color=COLOR_BORD)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(cell, texto, fill=None):
    cell.value = texto
    cell.font  = FONT_H
    cell.fill  = fill or FILL_DARK
    cell.alignment = ALIGN_C
    cell.border = BORDER


def cel(cell, valor, link=False, bold=False, wrap=False, zebra=None):
    cell.value = valor
    cell.font  = FONT_L if link else (FONT_BOLD if bold else FONT_B)
    cell.alignment = ALIGN_W if wrap else ALIGN_L
    cell.border = BORDER
    if zebra:
        cell.fill = zebra


# ---------------------------------------------------------------------------
def main():
    with open(DATA / "catalogo_profundo.json", encoding="utf-8") as f:
        productos = json.load(f)

    wb = Workbook()

    # -----------------------------------------------------------------------
    # HOJA 1 — Catálogo
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Catálogo"
    ws.freeze_panes = "A3"

    # Fila 1 — grupos
    n_specs = len(SPECS_COLUMNAS)
    grupos = [
        ("INFORMACIÓN BÁSICA",       1,           7,              FILL_DARK),
        ("COMERCIAL",                8,           9,              FILL_MID),
        ("CATEGORÍAS",               10,          12,             FILL_DARK),
        ("ESPECIFICACIONES TÉCNICAS",13,          12 + n_specs,   FILL_MID),
        ("DESCRIPCIÓN",              13 + n_specs, 13 + n_specs,  FILL_DARK),
    ]
    for etiqueta, c1, c2, fill in grupos:
        cell = ws.cell(row=1, column=c1)
        cell.value = etiqueta
        cell.font  = FONT_H
        cell.fill  = fill
        cell.alignment = ALIGN_C
        cell.border = BORDER
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1,
                           end_row=1,   end_column=c2)

    # Fila 2 — nombres de columna
    headers = (
        ["SKU", "Marca", "Título", "Precio CLP", "URL Producto", "URL Imagen", "Disponibilidad"]
        + ["EAN", "URLs Imagen (cantidad)"]
        + ["Categoría 1", "Categoría 2", "Categoría 3"]
        + SPECS_COLUMNAS
        + ["Descripción completa"]
    )
    for col, h in enumerate(headers, 1):
        if col <= 7:       fill = FILL_DARK
        elif col <= 9:     fill = FILL_MID
        elif col <= 12:    fill = FILL_DARK
        elif col <= 12 + n_specs: fill = FILL_MID
        else:              fill = FILL_DARK
        hdr(ws.cell(row=2, column=col), h, fill)

    # Filas de datos
    for fi, prod in enumerate(productos, 3):
        zfill = FILL_LIGHT if fi % 2 == 0 else None
        cats  = prod.get("categorias") or []
        specs = prod.get("especificaciones") or {}

        def c(col):
            cell = ws.cell(row=fi, column=col)
            if zfill:
                cell.fill = zfill
            return cell

        cel(c(1),  str(prod.get("sku", "")),  bold=True, zebra=zfill)
        cel(c(2),  prod.get("marca", ""),      zebra=zfill)
        cel(c(3),  prod.get("titulo", ""),     zebra=zfill)

        # Precio con formato numérico
        pc = c(4)
        pc.value = int(prod.get("precio_clp") or 0)
        pc.number_format = "#,##0"
        pc.font = FONT_B
        pc.alignment = ALIGN_C
        pc.border = BORDER

        cel(c(5),  prod.get("url", ""),        link=True,  zebra=zfill)
        cel(c(6),  prod.get("url_imagen", ""), link=True,  zebra=zfill)
        cel(c(7),  prod.get("disponibilidad", ""),         zebra=zfill)

        # EAN
        cel(c(8),  prod.get("ean", ""),        zebra=zfill)

        # Cantidad de imágenes en la galería
        n_imgs = len(prod.get("urls_imagen") or [])
        ci = c(9)
        ci.value = n_imgs if n_imgs else ""
        ci.font  = FONT_B
        ci.alignment = ALIGN_C
        ci.border = BORDER
        if zfill:
            ci.fill = zfill

        # Categorías
        for idx, cat in enumerate(cats[:3], 10):
            cel(c(idx), cat, zebra=zfill)

        # Specs
        for si, key in enumerate(SPECS_COLUMNAS, 13):
            cel(c(si), specs.get(key, ""), zebra=zfill)

        # Descripción (HTML → solo texto, recortado a 2000 chars)
        desc_raw = prod.get("descripcion_completa") or ""
        # Quitar tags HTML básicos
        import re
        desc_txt = re.sub(r"<[^>]+>", " ", desc_raw)
        desc_txt = re.sub(r"\s+", " ", desc_txt).strip()[:2000]
        cel(c(13 + n_specs), desc_txt, wrap=True, zebra=zfill)

    # Anchos
    anchos = {
        1: 14, 2: 18, 3: 55, 4: 14, 5: 45, 6: 45, 7: 15,
        8: 16, 9: 12,
        10: 28, 11: 28, 12: 28,
    }
    for i in range(13, 13 + n_specs):
        anchos[i] = 22
    anchos[13 + n_specs] = 60

    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    last_col = get_column_letter(13 + n_specs)
    ws.auto_filter.ref = f"A2:{last_col}2"

    # -----------------------------------------------------------------------
    # HOJA 2 — Resumen
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 20

    def h2(cell, texto):
        cell.value = texto
        cell.font  = FONT_H
        cell.fill  = FILL_DARK
        cell.alignment = ALIGN_C
        cell.border = BORDER

    def v2(cell, valor, fmt=None, fi=None):
        cell.value = valor
        cell.font  = FONT_B
        cell.border = BORDER
        cell.alignment = ALIGN_L
        if fmt:
            cell.number_format = fmt
        if fi and fi % 2 == 0:
            cell.fill = FILL_LIGHT

    n = len(productos)
    h2(ws2.cell(row=1, column=1), "Métrica")
    h2(ws2.cell(row=1, column=2), "Valor")
    ws2.row_dimensions[1].height = 24

    filas_resumen = [
        ("Total productos",              f"=COUNTA('Catálogo'!A3:A{n+2})"),
        ("Marcas únicas",                f"=SUMPRODUCT(1/COUNTIF('Catálogo'!B3:B{n+2},'Catálogo'!B3:B{n+2}))"),
        ("Precio promedio (CLP)",        f"=AVERAGE('Catálogo'!D3:D{n+2})"),
        ("Precio mínimo (CLP)",          f"=MIN('Catálogo'!D3:D{n+2})"),
        ("Precio máximo (CLP)",          f"=MAX('Catálogo'!D3:D{n+2})"),
        ("Productos InStock",            f"=COUNTIF('Catálogo'!G3:G{n+2},\"InStock\")"),
        ("Productos OutOfStock",         f"=COUNTIF('Catálogo'!G3:G{n+2},\"OutOfStock\")"),
        ("Productos con EAN",            f"=COUNTA('Catálogo'!H3:H{n+2})"),
        ("Productos con galería (>1 img)",f"=COUNTIF('Catálogo'!I3:I{n+2},\">1\")"),
        ("Total especificaciones únicas","158"),
    ]

    for fi, (etiqueta, valor) in enumerate(filas_resumen, 2):
        ca = ws2.cell(row=fi, column=1)
        cb = ws2.cell(row=fi, column=2)
        ca.value = etiqueta
        ca.font  = FONT_B
        ca.border = BORDER
        ca.alignment = ALIGN_L
        cb.value = valor
        cb.font  = FONT_B
        cb.border = BORDER
        cb.alignment = ALIGN_L
        if "precio" in etiqueta.lower():
            cb.number_format = "#,##0"
        if fi % 2 == 0:
            ca.fill = FILL_LIGHT
            cb.fill = FILL_LIGHT

    # -----------------------------------------------------------------------
    wb.save(SALIDA)
    print(f"Excel guardado: {SALIDA}")
    print(f"Productos exportados: {len(productos)}")
    print(f"Columnas: {len(headers)}")


if __name__ == "__main__":
    main()
