"""Exporta weak_labels.json a un Excel navegable para revisión manual.

Hojas:
  1. Resumen         → conteos, ratios, distribución por cluster/tier/marca
  2. Positivos       → todos los positivos con colores por tier
  3. Negativos_duros → negativos intra-cluster (los más valiosos)
  4. Negativos_fac   → muestra de negativos inter-cluster
  5. Muestra_500     → mezcla balanceada para inspección rápida

Colores:
  - Positivos tier 1 (marca+dims)        → verde suave
  - Positivos tier 2 (marca+cosine+jacc) → verde medio
  - Positivos tier 3 (solo cosine+jacc)  → verde oscuro
  - Negativos duros                      → rojo suave
  - Negativos fáciles                    → gris
"""
from __future__ import annotations

import json
import random
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data")
IN_LABELS = BASE / "weak_labels.json"
IN_REPORTE = BASE / "weak_labels_reporte.json"
OUT_XLSX = BASE / "weak_labels_visualizacion.xlsx"

# Colores (Arial blanco como font principal)
COLOR_POS_T1   = "C8E6C9"  # verde suave
COLOR_POS_T2   = "81C784"  # verde medio
COLOR_POS_T3   = "388E3C"  # verde oscuro (font blanco)
COLOR_NEG_DURO = "FFCDD2"  # rojo suave
COLOR_NEG_FAC  = "ECEFF1"  # gris suave
COLOR_HEADER   = "37474F"  # azul grisáceo oscuro
COLOR_ZEBRA    = "F5F5F5"  # gris muy claro

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_POS_T3  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
FONT_TITLE   = Font(name="Arial", size=14, bold=True)

THIN = Side(border_style="thin", color="BDBDBD")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row: int, headers: list[str]):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 28


def _format_data_row(ws, row: int, values: list, fill_color: str | None = None,
                     font: Font = FONT_DEFAULT, zebra: bool = False):
    fill = None
    if fill_color:
        fill = PatternFill("solid", fgColor=fill_color)
    elif zebra:
        fill = PatternFill("solid", fgColor=COLOR_ZEBRA)
    for col, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = font
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER_ALL
        if fill:
            c.fill = fill


def _hoja_resumen(wb: Workbook, labels: list, reporte: dict):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

    r = 1
    ws.cell(row=r, column=1, value="WEAK LABELS - RESUMEN EJECUTIVO").font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 22
    r += 2

    stats = [
        ("Total pares etiquetados", reporte["total_pares"]),
        ("  Positivos (label=1)",    reporte["positivos"]),
        ("  Negativos duros",        reporte["negativos_duros"]),
        ("  Negativos fáciles",      reporte["negativos_faciles"]),
        ("Ratio pos:neg",            f"1 : {round((reporte['negativos_duros']+reporte['negativos_faciles'])/max(1,reporte['positivos']),2)}"),
        ("Pares intra-cluster totales", reporte["pares_intra_cluster_total"]),
        ("Pares descartados (dudosos)", reporte["pares_descartados"]),
    ]
    for k, v in stats:
        ws.cell(row=r, column=1, value=k).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=2, value=v).font = FONT_DEFAULT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="PARÁMETROS DE ETIQUETADO").font = Font(name="Arial", size=12, bold=True)
    r += 1
    for k, v in reporte["parametros"].items():
        ws.cell(row=r, column=1, value=f"  {k}").font = FONT_DEFAULT
        ws.cell(row=r, column=2, value=v).font = FONT_DEFAULT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="POSITIVOS POR CLUSTER").font = Font(name="Arial", size=12, bold=True)
    r += 1
    _header_row(ws, r, ["cluster_id", "n_positivos", "ejemplo_titulo"])
    r += 1
    # ejemplo por cluster
    por_cluster = {}
    for rec in labels:
        if rec["label"] == 1:
            cid = rec["ctx"]["cluster_id"]
            if cid not in por_cluster:
                por_cluster[cid] = {"n": 0, "ej": rec["ctx"]["easy_titulo"]}
            por_cluster[cid]["n"] += 1
    for cid, info in sorted(por_cluster.items(), key=lambda kv: -kv[1]["n"]):
        _format_data_row(ws, r, [cid, info["n"], info["ej"][:80]])
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="TOP MARCAS EN POSITIVOS").font = Font(name="Arial", size=12, bold=True)
    r += 1
    _header_row(ws, r, ["marca_norm", "n_positivos", "%"])
    r += 1
    total_pos = reporte["positivos"]
    for marca, n in reporte["marcas_top_en_positivos"].items():
        pct = f"{n*100/total_pos:.1f}%"
        _format_data_row(ws, r, [marca or "(vacía)", n, pct])
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="FUENTE DE ETIQUETAS").font = Font(name="Arial", size=12, bold=True)
    r += 1
    _header_row(ws, r, ["fuente", "n_pares", "label"])
    r += 1
    fuentes = Counter((rec["ctx"]["fuente"], rec["label"]) for rec in labels)
    for (fuente, label), n in sorted(fuentes.items(), key=lambda kv: -kv[1]):
        _format_data_row(ws, r, [fuente, n, label])
        r += 1


COLS_DETALLE = [
    "label", "fuente", "cluster", "cosine",
    "Δlargo", "Δdiam", "brand", "mat", "rosca", "cabeza",
    "jaccard", "ratio_p",
    "easy_marca", "easy_título",
    "sodimac_marca", "sodimac_título",
    "easy_sku", "sodimac_sku",
]
WIDTHS_DETALLE = [8, 18, 10, 9, 9, 9, 8, 6, 7, 8, 10, 10, 14, 55, 14, 55, 14, 14]


def _fila_detalle(rec: dict) -> list:
    f = rec["features"]
    ctx = rec["ctx"]
    return [
        rec["label"],
        ctx.get("fuente", ""),
        ctx.get("cluster_id", -1),
        f.get("cosine_sim"),
        f.get("diff_largo_mm"),
        f.get("diff_diametro_mm"),
        f.get("brand_match"),
        f.get("material_match"),
        f.get("tipo_rosca_match"),
        f.get("tipo_cabeza_match"),
        f.get("jaccard_titulo"),
        f.get("ratio_precio"),
        ctx.get("easy_marca") or "",
        ctx.get("easy_titulo") or "",
        ctx.get("sodimac_marca") or "",
        ctx.get("sodimac_titulo") or "",
        ctx.get("easy_sku") or "",
        ctx.get("sodimac_sku") or "",
    ]


def _color_pos_por_tier(tier_counts_por_rec: dict, rec: dict) -> tuple[str, Font]:
    """Recibe mapa rec_id → tier (None si no se conoce). Devuelve (color, font)."""
    # recuperamos tier mirando si features tiene _tier o mediante lookup
    t = rec.get("_tier_cached")
    if t == "tier1_marca_dims":
        return COLOR_POS_T1, FONT_DEFAULT
    if t == "tier2_marca_cosine_jaccard":
        return COLOR_POS_T2, FONT_DEFAULT
    if t == "tier3_cosine_jaccard":
        return COLOR_POS_T3, FONT_POS_T3
    return COLOR_POS_T1, FONT_DEFAULT  # fallback verde suave


def _hoja_positivos(wb: Workbook, labels: list):
    positivos = [r for r in labels if r["label"] == 1]
    ws = wb.create_sheet("Positivos")
    _set_col_widths(ws, WIDTHS_DETALLE)
    _header_row(ws, 1, COLS_DETALLE)
    ws.freeze_panes = "A2"

    for i, rec in enumerate(positivos, start=2):
        # inferir tier por los features (regeneramos la lógica rápido)
        f = rec["features"]
        brand = f.get("brand_match", 0)
        cos = f.get("cosine_sim", 0)
        jac = f.get("jaccard_titulo", 0)
        dl = f.get("diff_largo_mm")
        dd = f.get("diff_diametro_mm")
        if brand and cos >= 0.65 and (dl is not None or dd is not None):
            color = COLOR_POS_T1; font = FONT_DEFAULT
        elif brand and cos >= 0.82 and jac >= 0.30:
            color = COLOR_POS_T2; font = FONT_DEFAULT
        elif cos >= 0.88 and jac >= 0.45:
            color = COLOR_POS_T3; font = FONT_POS_T3
        else:
            color = COLOR_POS_T1; font = FONT_DEFAULT

        _format_data_row(ws, i, _fila_detalle(rec), fill_color=color, font=font)


def _hoja_negativos_duros(wb: Workbook, labels: list):
    neg = [r for r in labels if r["label"] == 0 and r["ctx"].get("fuente") == "neg_duro_cluster"]
    # ordenar por cosine_sim descendente (los más "confusos" primero)
    neg.sort(key=lambda r: -(r["features"].get("cosine_sim") or 0))
    ws = wb.create_sheet("Negativos_duros")
    _set_col_widths(ws, WIDTHS_DETALLE)
    _header_row(ws, 1, COLS_DETALLE)
    ws.freeze_panes = "A2"
    for i, rec in enumerate(neg, start=2):
        _format_data_row(ws, i, _fila_detalle(rec), fill_color=COLOR_NEG_DURO)


def _hoja_negativos_faciles(wb: Workbook, labels: list, max_rows: int = 300):
    neg = [r for r in labels if r["label"] == 0 and r["ctx"].get("fuente") == "neg_facil_inter"]
    neg = neg[:max_rows]  # limitar para no inflar el Excel
    ws = wb.create_sheet("Negativos_fac")
    _set_col_widths(ws, WIDTHS_DETALLE)
    _header_row(ws, 1, COLS_DETALLE)
    ws.freeze_panes = "A2"
    for i, rec in enumerate(neg, start=2):
        _format_data_row(ws, i, _fila_detalle(rec), fill_color=COLOR_NEG_FAC)


def _hoja_muestra_balanceada(wb: Workbook, labels: list):
    """500 filas: todos los positivos + muestra negativos, mezclados."""
    random.seed(7)
    pos = [r for r in labels if r["label"] == 1]
    neg_d = [r for r in labels if r["label"] == 0 and r["ctx"].get("fuente") == "neg_duro_cluster"]
    neg_f = [r for r in labels if r["label"] == 0 and r["ctx"].get("fuente") == "neg_facil_inter"]
    random.shuffle(neg_d)
    random.shuffle(neg_f)
    muestra = pos + neg_d[:len(pos)*2] + neg_f[:len(pos)]
    random.shuffle(muestra)

    ws = wb.create_sheet("Muestra_mezclada")
    _set_col_widths(ws, WIDTHS_DETALLE)
    _header_row(ws, 1, COLS_DETALLE)
    ws.freeze_panes = "A2"

    for i, rec in enumerate(muestra, start=2):
        if rec["label"] == 1:
            color = COLOR_POS_T1
            font = FONT_DEFAULT
        elif rec["ctx"].get("fuente") == "neg_duro_cluster":
            color = COLOR_NEG_DURO
            font = FONT_DEFAULT
        else:
            color = COLOR_NEG_FAC
            font = FONT_DEFAULT
        _format_data_row(ws, i, _fila_detalle(rec), fill_color=color, font=font)


def main():
    print("=" * 70)
    print("EXPORTANDO WEAK LABELS A EXCEL")
    print("=" * 70)

    with open(IN_LABELS, encoding="utf-8") as f:
        labels = json.load(f)
    with open(IN_REPORTE, encoding="utf-8") as f:
        reporte = json.load(f)

    print(f"  Total pares cargados : {len(labels)}")
    print(f"  Positivos            : {sum(1 for r in labels if r['label']==1)}")
    print(f"  Negativos            : {sum(1 for r in labels if r['label']==0)}")

    wb = Workbook()
    # la default sheet la reutilizamos para Resumen vía _hoja_resumen (index 0)
    # primero eliminamos la default
    default = wb.active
    wb.remove(default)

    _hoja_resumen(wb, labels, reporte)
    _hoja_positivos(wb, labels)
    _hoja_negativos_duros(wb, labels)
    _hoja_negativos_faciles(wb, labels)
    _hoja_muestra_balanceada(wb, labels)

    wb.save(OUT_XLSX)
    size_kb = OUT_XLSX.stat().st_size // 1024
    print(f"\n  Excel guardado: {OUT_XLSX}")
    print(f"  Tamaño: {size_kb} KB")
    print(f"  Hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
