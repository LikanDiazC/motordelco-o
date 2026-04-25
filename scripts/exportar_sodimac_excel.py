"""Exporta catalogo_profundo.json de Sodimac a Excel con formato profesional."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

DATA  = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data\sodimac")
SALIDA = DATA / "catalogo_sodimac.xlsx"

# Specs que merecen columna propia (top más comunes y útiles)
SPECS_COLUMNAS = [
    "Largo",
    "Diámetro",
    "Material",
    "Tipo de cabeza",
    "Tipo de tornillo",
    "Tipo de perno",
    "Color",
    "Cantidad por paquete",
    "Superficie de aplicación",
    "País de origen",
    "Garantía",
    "Detalle de la garantía",
    "Modelo",
    "Presentación",
    "Características",
]

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
COLOR_HEADER_DARK  = "1F3864"   # azul marino
COLOR_HEADER_MID   = "2E75B6"   # azul medio
COLOR_HEADER_LIGHT = "D9E1F2"   # azul muy claro (zebra par)
COLOR_WHITE        = "FFFFFF"
COLOR_BORDER       = "B8CCE4"

FONT_HEADER = Font(name="Arial", bold=True, color=COLOR_WHITE, size=10)
FONT_BODY   = Font(name="Arial", size=9)
FONT_LINK   = Font(name="Arial", size=9, color="0563C1", underline="single")
FONT_BOLD   = Font(name="Arial", size=9, bold=True)

FILL_DARK   = PatternFill("solid", fgColor=COLOR_HEADER_DARK)
FILL_MID    = PatternFill("solid", fgColor=COLOR_HEADER_MID)
FILL_LIGHT  = PatternFill("solid", fgColor=COLOR_HEADER_LIGHT)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
ALIGN_WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

thin = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_header(cell, texto, fill=None):
    cell.value = texto
    cell.font  = FONT_HEADER
    cell.fill  = fill or FILL_DARK
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER


def aplicar_celda(cell, valor, link=False, bold=False, wrap=False):
    cell.value = valor
    if link:
        cell.font = FONT_LINK
    elif bold:
        cell.font = FONT_BOLD
    else:
        cell.font = FONT_BODY
    cell.alignment = ALIGN_WRAP if wrap else ALIGN_LEFT
    cell.border = BORDER


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    with open(DATA / "catalogo_profundo.json", encoding="utf-8") as f:
        productos = json.load(f)

    wb = Workbook()

    # -----------------------------------------------------------------------
    # HOJA 1 — Catálogo completo
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Catálogo"
    ws.freeze_panes = "A3"

    # Fila 1: grupos de encabezado
    grupos = [
        ("INFORMACIÓN BÁSICA", 1, 7, FILL_DARK),
        ("COMERCIAL", 8, 10, FILL_MID),
        ("CATEGORÍAS", 11, 13, FILL_DARK),
        ("ESPECIFICACIONES TÉCNICAS", 14, 14 + len(SPECS_COLUMNAS) - 1, FILL_MID),
        ("DESCRIPCIÓN", 14 + len(SPECS_COLUMNAS), 14 + len(SPECS_COLUMNAS), FILL_DARK),
    ]
    for etiqueta, col_ini, col_fin, fill in grupos:
        c = ws.cell(row=1, column=col_ini)
        c.value = etiqueta
        c.font  = FONT_HEADER
        c.fill  = fill
        c.alignment = ALIGN_CENTER
        c.border = BORDER
        if col_fin > col_ini:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=1,   end_column=col_fin
            )

    # Fila 2: encabezados de columna
    headers_basicos = ["SKU", "Marca", "Título", "Precio CLP", "URL Producto", "URL Imagen", "Disponibilidad"]
    headers_comercial = ["Rating", "Reseñas", "Precio c/Desc."]
    headers_cats = ["Categoría 1", "Categoría 2", "Categoría 3"]
    headers_desc = ["Descripción"]

    all_headers = (headers_basicos + headers_comercial + headers_cats
                   + SPECS_COLUMNAS + headers_desc)

    for col, h in enumerate(all_headers, 1):
        if col <= 7:
            fill = FILL_DARK
        elif col <= 10:
            fill = FILL_MID
        elif col <= 13:
            fill = FILL_DARK
        elif col <= 13 + len(SPECS_COLUMNAS):
            fill = FILL_MID
        else:
            fill = FILL_DARK
        aplicar_header(ws.cell(row=2, column=col), h, fill)

    # Datos
    for fila_idx, prod in enumerate(productos, 3):
        zebra = FILL_LIGHT if fila_idx % 2 == 0 else None
        cats  = prod.get("categorias") or []
        specs = prod.get("especificaciones") or {}

        def c(col):
            cell = ws.cell(row=fila_idx, column=col)
            if zebra:
                cell.fill = zebra
            return cell

        aplicar_celda(c(1),  str(prod.get("sku", "")),   bold=True)
        aplicar_celda(c(2),  prod.get("marca", ""))
        aplicar_celda(c(3),  prod.get("titulo", ""))
        precio = prod.get("precio_clp") or 0
        cell_precio = c(4)
        cell_precio.value = int(precio) if precio else 0
        cell_precio.number_format = '#,##0'
        cell_precio.font = FONT_BODY
        cell_precio.alignment = ALIGN_CENTER
        cell_precio.border = BORDER
        if zebra:
            cell_precio.fill = zebra
        aplicar_celda(c(5),  prod.get("url", ""),         link=True)
        aplicar_celda(c(6),  prod.get("url_imagen", ""),  link=True)
        aplicar_celda(c(7),  prod.get("disponibilidad", ""))

        # Comercial
        rating = prod.get("rating")
        cell_r = c(8)
        cell_r.value = rating
        cell_r.number_format = '0.0'
        cell_r.font = FONT_BODY
        cell_r.alignment = ALIGN_CENTER
        cell_r.border = BORDER
        if zebra:
            cell_r.fill = zebra

        reviews = prod.get("review_count")
        cell_rv = c(9)
        cell_rv.value = reviews
        cell_rv.number_format = '#,##0'
        cell_rv.font = FONT_BODY
        cell_rv.alignment = ALIGN_CENTER
        cell_rv.border = BORDER
        if zebra:
            cell_rv.fill = zebra

        # precio con descuento (placeholder = mismo precio, se calcula si hay normal)
        aplicar_celda(c(10), "")

        # Categorías
        for ci, cat_val in enumerate(cats[:3], 11):
            aplicar_celda(c(ci), cat_val)

        # Specs técnicas
        for si, spec_key in enumerate(SPECS_COLUMNAS, 14):
            aplicar_celda(c(si), specs.get(spec_key, ""))

        # Descripción
        desc_col = 14 + len(SPECS_COLUMNAS)
        aplicar_celda(c(desc_col), prod.get("descripcion", ""), wrap=True)

    # Anchos de columna
    anchos = {
        1: 14,   # SKU
        2: 16,   # Marca
        3: 55,   # Título
        4: 14,   # Precio
        5: 45,   # URL
        6: 45,   # Imagen
        7: 14,   # Disponibilidad
        8: 9,    # Rating
        9: 10,   # Reseñas
        10: 12,  # Precio desc
        11: 28,  # Cat 1
        12: 28,  # Cat 2
        13: 28,  # Cat 3
    }
    for i, spec_key in enumerate(SPECS_COLUMNAS, 14):
        anchos[i] = 22
    anchos[14 + len(SPECS_COLUMNAS)] = 60   # Descripción

    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # Filtros automáticos desde fila 2
    ultima_col = get_column_letter(14 + len(SPECS_COLUMNAS))
    ws.auto_filter.ref = f"A2:{ultima_col}2"

    # -----------------------------------------------------------------------
    # HOJA 2 — Resumen estadístico
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet("Resumen")

    def h2(cell, texto):
        cell.value = texto
        cell.font  = Font(name="Arial", bold=True, color=COLOR_WHITE, size=10)
        cell.fill  = FILL_DARK
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER

    def v2(cell, valor, fmt=None):
        cell.value = valor
        cell.font  = FONT_BODY
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER
        if fmt:
            cell.number_format = fmt

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18

    metricas = [
        ("Métrica", "Valor"),
        ("Total productos", f"=COUNTA('Catálogo'!A3:A{len(productos)+2})"),
        ("Marcas únicas", f"=SUMPRODUCT(1/COUNTIF('Catálogo'!B3:B{len(productos)+2},'Catálogo'!B3:B{len(productos)+2}))"),
        ("Precio promedio (CLP)", f"=AVERAGE('Catálogo'!D3:D{len(productos)+2})"),
        ("Precio mínimo (CLP)", f"=MIN('Catálogo'!D3:D{len(productos)+2})"),
        ("Precio máximo (CLP)", f"=MAX('Catálogo'!D3:D{len(productos)+2})"),
        ("Productos con rating", f"=COUNTA('Catálogo'!H3:H{len(productos)+2})"),
        ("Rating promedio", f"=AVERAGEIF('Catálogo'!H3:H{len(productos)+2},\">0\",'Catálogo'!H3:H{len(productos)+2})"),
        ("Productos InStock", f"=COUNTIF('Catálogo'!G3:G{len(productos)+2},\"InStock\")"),
        ("Productos sin disponibilidad", f"=COUNTBLANK('Catálogo'!G3:G{len(productos)+2})"),
    ]

    for fila, (etiqueta, valor) in enumerate(metricas, 1):
        if fila == 1:
            h2(ws2.cell(row=1, column=1), etiqueta)
            h2(ws2.cell(row=1, column=2), str(valor))
        else:
            v2(ws2.cell(row=fila, column=1), etiqueta)
            cell_v = ws2.cell(row=fila, column=2)
            cell_v.value = valor
            cell_v.font  = FONT_BODY
            cell_v.border = BORDER
            if "precio" in etiqueta.lower() or "precio" in etiqueta.lower():
                cell_v.number_format = "#,##0"
            elif "rating" in etiqueta.lower():
                cell_v.number_format = "0.00"
            cell_v.alignment = ALIGN_LEFT
            # zebra
            if fila % 2 == 0:
                ws2.cell(row=fila, column=1).fill = FILL_LIGHT
                cell_v.fill = FILL_LIGHT

    ws2.row_dimensions[1].height = 24

    # -----------------------------------------------------------------------
    # Guardar
    # -----------------------------------------------------------------------
    wb.save(SALIDA)
    print(f"Excel guardado: {SALIDA}")
    print(f"Productos exportados: {len(productos)}")
    print(f"Columnas por producto: {len(all_headers)}")


if __name__ == "__main__":
    main()
