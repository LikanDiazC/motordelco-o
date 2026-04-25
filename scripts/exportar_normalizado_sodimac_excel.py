"""Exporta catalogo_normalizado.json de Sodimac a Excel."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA   = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data\sodimac")
SALIDA = DATA / "catalogo_normalizado_sodimac.xlsx"

# ---------------------------------------------------------------------------
# Paleta Sodimac (naranja corporativo + azul oscuro)
# ---------------------------------------------------------------------------
C_DARK  = "1F3864"
C_MID   = "2E75B6"
C_LIGHT = "D9E1F2"
C_ORG   = "C0392B"   # rojo-naranja para rating (destaca)
C_ORLT  = "FADBD8"
C_GREEN = "196F3D"
C_GRLT  = "D5F5E3"
C_BORD  = "B8CCE4"

FONT_H    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
FONT_B    = Font(name="Arial", size=9)
FONT_BOLD = Font(name="Arial", size=9, bold=True)
FONT_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")
FONT_NUM  = Font(name="Arial", size=9, color="1A5276")
FONT_RTG  = Font(name="Arial", size=9, color=C_ORG, bold=True)

FILL_DARK  = PatternFill("solid", fgColor=C_DARK)
FILL_MID   = PatternFill("solid", fgColor=C_MID)
FILL_LIGHT = PatternFill("solid", fgColor=C_LIGHT)
FILL_ORG   = PatternFill("solid", fgColor=C_ORG)
FILL_ORLT  = PatternFill("solid", fgColor=C_ORLT)
FILL_GREEN = PatternFill("solid", fgColor=C_GREEN)
FILL_GRLT  = PatternFill("solid", fgColor=C_GRLT)

AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left",   vertical="center")
AL_W = Alignment(horizontal="left",   vertical="top", wrap_text=True)

thin   = Side(style="thin", color=C_BORD)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, row, col, texto, fill=None, span=None):
    cell = ws.cell(row=row, column=col)
    cell.value = texto
    cell.font  = FONT_H
    cell.fill  = fill or FILL_DARK
    cell.alignment = AL_C
    cell.border = BORDER
    if span and span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)


def cel(ws, row, col, valor, fmt=None, link=False, bold=False,
        wrap=False, zfill=None, center=False, rating=False):
    cell = ws.cell(row=row, column=col)
    cell.value = valor
    if rating:
        cell.font = FONT_RTG
    elif link:
        cell.font = FONT_LINK
    elif bold:
        cell.font = FONT_BOLD
    elif fmt:
        cell.font = FONT_NUM
    else:
        cell.font = FONT_B
    cell.alignment = AL_W if wrap else (AL_C if center else AL_L)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if zfill:
        cell.fill = zfill


# ---------------------------------------------------------------------------
def main():
    with open(DATA / "catalogo_normalizado.json", encoding="utf-8") as f:
        data = json.load(f)

    n = len(data)
    wb = Workbook()

    # =======================================================================
    # HOJA 1 — Catálogo normalizado
    # =======================================================================
    ws = wb.active
    ws.title = "Catálogo normalizado"
    ws.freeze_panes = "A3"

    # Fila 1 — grupos
    grupos = [
        ("ID & BÁSICO",          1,  7,  FILL_DARK),
        ("PRECIOS",              8, 10,  FILL_MID),
        ("RATING",              11, 12,  FILL_ORG),
        ("CATEGORÍAS",          13, 16,  FILL_DARK),
        ("DIMENSIONES",         17, 19,  FILL_GREEN),
        ("ATRIBUTOS TÉCNICOS",  20, 25,  FILL_MID),
        ("DISPONIBILIDAD",      26, 26,  FILL_DARK),
    ]
    for etiqueta, c1, c2, fill in grupos:
        hdr(ws, 1, c1, etiqueta, fill, span=c2 - c1 + 1)

    # Fila 2 — encabezados
    headers = [
        # ID & BÁSICO (1-7)
        "ID Producto", "Tienda", "SKU", "Marca", "Título",
        "URL Producto", "URL Imagen",
        # PRECIOS (8-10)
        "Precio CLP", "Precio Normal CLP", "Descuento %",
        # RATING (11-12)
        "Rating", "Reseñas",
        # CATEGORÍAS (13-16)
        "Categoría 1", "Categoría 2", "Categoría 3", "Categoría 4",
        # DIMENSIONES (17-19)
        "Largo (mm)", "Diámetro (mm)", "Medida cruda",
        # ATRIBUTOS TÉCNICOS (20-25)
        "Material", "Tipo cabeza", "Tipo rosca", "Tipo punta",
        "Color", "Cantidad empaque",
        # DISPONIBILIDAD (26)
        "Disponibilidad",
    ]

    fills_h2 = (
        [FILL_DARK] * 7 + [FILL_MID] * 3 + [FILL_ORG] * 2
        + [FILL_DARK] * 4 + [FILL_GREEN] * 3 + [FILL_MID] * 6
        + [FILL_DARK]
    )
    for col, (h, fill) in enumerate(zip(headers, fills_h2), 1):
        hdr(ws, 2, col, h, fill)

    # Filas de datos
    for fi, prod in enumerate(data, 3):
        zf     = FILL_LIGHT if fi % 2 == 0 else None
        zf_dim = FILL_GRLT  if fi % 2 == 0 else None
        zf_rtg = FILL_ORLT  if fi % 2 == 0 else None

        mb   = prod.get("metadata_basica") or {}
        mt   = prod.get("metadata_tecnica") or {}
        dims = mt.get("dimensiones") or {}
        cats = mb.get("categorias") or []

        precio_actual = mb.get("precio_clp") or 0
        precio_normal = mb.get("precio_normal_clp") or precio_actual
        desc_pct      = mb.get("descuento_pct")
        rating        = mb.get("rating")
        reviews       = mb.get("review_count")

        # ID & BÁSICO
        cel(ws, fi, 1, prod.get("id_producto", ""),   bold=True,  zfill=zf)
        cel(ws, fi, 2, prod.get("tienda", ""),         center=True, zfill=zf)
        cel(ws, fi, 3, str(prod.get("sku", "")),       bold=True,  zfill=zf)
        cel(ws, fi, 4, mb.get("marca", ""),             zfill=zf)
        cel(ws, fi, 5, mb.get("titulo", ""),            zfill=zf)
        cel(ws, fi, 6, prod.get("url_producto", ""),   link=True,  zfill=zf)
        cel(ws, fi, 7, prod.get("url_imagen", ""),     link=True,  zfill=zf)

        # PRECIOS
        for col_p, val_p in [(8, precio_actual), (9, precio_normal)]:
            cp = ws.cell(row=fi, column=col_p)
            cp.value = int(val_p) if val_p else 0
            cp.number_format = "#,##0"
            cp.font = FONT_NUM
            cp.alignment = AL_C
            cp.border = BORDER
            if zf: cp.fill = zf

        cd = ws.cell(row=fi, column=10)
        cd.value = desc_pct if desc_pct else ""
        cd.number_format = '0.0"%"'
        cd.font = FONT_NUM
        cd.alignment = AL_C
        cd.border = BORDER
        if zf: cd.fill = zf

        # RATING
        cr = ws.cell(row=fi, column=11)
        cr.value = rating if rating else ""
        cr.number_format = "0.0"
        cr.font = FONT_RTG if rating else FONT_B
        cr.alignment = AL_C
        cr.border = BORDER
        if zf_rtg: cr.fill = zf_rtg

        crv = ws.cell(row=fi, column=12)
        crv.value = reviews if reviews else ""
        crv.number_format = "#,##0"
        crv.font = FONT_RTG if reviews else FONT_B
        crv.alignment = AL_C
        crv.border = BORDER
        if zf_rtg: crv.fill = zf_rtg

        # CATEGORÍAS
        for ci, cat in enumerate(cats[:4], 13):
            cel(ws, fi, ci, cat, zfill=zf)
        for ci in range(len(cats[:4]) + 13, 17):
            cel(ws, fi, ci, "", zfill=zf)

        # DIMENSIONES
        largo = dims.get("largo_mm")
        diam  = dims.get("diametro_mm")

        for col_d, val_d in [(17, largo), (18, diam)]:
            cd2 = ws.cell(row=fi, column=col_d)
            cd2.value = val_d if val_d else ""
            cd2.number_format = "0.00"
            cd2.font = FONT_NUM
            cd2.alignment = AL_C
            cd2.border = BORDER
            if zf_dim: cd2.fill = zf_dim

        cel(ws, fi, 19, dims.get("medida_cruda", ""), zfill=zf)

        # ATRIBUTOS TÉCNICOS
        cel(ws, fi, 20, mt.get("material", ""),    zfill=zf)
        cel(ws, fi, 21, mt.get("tipo_cabeza", ""), zfill=zf)
        cel(ws, fi, 22, mt.get("tipo_rosca", ""),  zfill=zf)
        cel(ws, fi, 23, mt.get("tipo_punta", ""),  zfill=zf)
        cel(ws, fi, 24, mt.get("color", ""),       zfill=zf)

        cc = ws.cell(row=fi, column=25)
        cc.value = int(mt.get("cantidad_empaque") or 1)
        cc.number_format = "#,##0"
        cc.font = FONT_NUM
        cc.alignment = AL_C
        cc.border = BORDER
        if zf: cc.fill = zf

        # DISPONIBILIDAD
        cel(ws, fi, 26, mb.get("disponibilidad", ""), center=True, zfill=zf)

    # Anchos de columna
    anchos = {
        1: 26, 2: 10, 3: 14, 4: 16, 5: 52,
        6: 42, 7: 42,
        8: 15, 9: 17, 10: 13,
        11: 9, 12: 10,
        13: 28, 14: 28, 15: 28, 16: 28,
        17: 12, 18: 14, 19: 18,
        20: 14, 21: 14, 22: 14, 23: 14, 24: 14, 25: 16,
        26: 14,
    }
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 34

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # =======================================================================
    # HOJA 2 — Resumen
    # =======================================================================
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 20

    hdr(ws2, 1, 1, "Métrica", FILL_DARK)
    hdr(ws2, 1, 2, "Valor",   FILL_DARK)
    ws2.row_dimensions[1].height = 24

    sh = "'Catálogo normalizado'"
    metricas = [
        ("Total productos",                    f"=COUNTA({sh}!C3:C{n+2})"),
        ("Marcas únicas",                      f"=SUMPRODUCT(1/COUNTIF({sh}!D3:D{n+2},{sh}!D3:D{n+2}))"),
        ("Precio promedio (CLP)",              f"=AVERAGE({sh}!H3:H{n+2})"),
        ("Precio mínimo (CLP)",                f"=MIN({sh}!H3:H{n+2})"),
        ("Precio máximo (CLP)",                f"=MAX({sh}!H3:H{n+2})"),
        ("Productos con rating",               f"=COUNTA({sh}!K3:K{n+2})"),
        ("Rating promedio",                    f"=AVERAGEIF({sh}!K3:K{n+2},\">0\",{sh}!K3:K{n+2})"),
        ("Productos InStock",                  f"=COUNTIF({sh}!Z3:Z{n+2},\"InStock\")"),
        ("Productos OutOfStock",               f"=COUNTIF({sh}!Z3:Z{n+2},\"OutOfStock\")"),
        ("Productos con largo_mm extraído",    f"=COUNTA({sh}!Q3:Q{n+2})"),
        ("Productos con diámetro_mm extraído", f"=COUNTA({sh}!R3:R{n+2})"),
        ("Productos con material",             f"=COUNTA({sh}!T3:T{n+2})"),
        ("Productos con categorías",           f"=COUNTA({sh}!M3:M{n+2})"),
        ("Cantidad promedio por empaque",      f"=AVERAGE({sh}!Y3:Y{n+2})"),
    ]

    for fi, (etiqueta, valor) in enumerate(metricas, 2):
        ca = ws2.cell(row=fi, column=1)
        cb = ws2.cell(row=fi, column=2)
        ca.value = etiqueta
        ca.font  = FONT_B
        ca.border = BORDER
        ca.alignment = AL_L
        cb.value = valor
        cb.font  = FONT_NUM
        cb.border = BORDER
        cb.alignment = AL_L
        if any(k in etiqueta.lower() for k in ("precio", "promedio", "cantidad")):
            cb.number_format = "#,##0"
        if "rating" in etiqueta.lower() and "promedio" in etiqueta.lower():
            cb.number_format = "0.00"
        if fi % 2 == 0:
            ca.fill = FILL_LIGHT
            cb.fill = FILL_LIGHT

    # =======================================================================
    wb.save(SALIDA)
    print(f"Excel guardado: {SALIDA}")
    print(f"Productos: {n} | Columnas: {len(headers)}")


if __name__ == "__main__":
    main()
