"""Exporta matches V2 a Excel de revisión humana, con comparación V1 vs V2."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data")
IN_V2_CANDS = BASE / "matches_v2_candidatos.json"
IN_V2_TOP   = BASE / "matches_v2_top.json"
IN_V1_TOP   = BASE / "matches_top.json"
IN_COMPARE  = BASE / "matches_v2_vs_v1.json"
OUT_XLSX    = BASE / "matches_v2_revisar.xlsx"

COLOR_HEADER  = "37474F"
COLOR_MATCH   = "C8E6C9"
COLOR_NEW_V2  = "FFE082"   # ámbar para matches solo en V2
COLOR_SOLO_V1 = "EF9A9A"   # rojo suave para los que V2 rechazó
COLOR_AMB     = "FFE0B2"
COLOR_ALTA    = "E8F5E9"
COLOR_MEDIA   = "FFF9C4"

FONT_DEFAULT = Font(name="Arial", size=10)
FONT_HEADER  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_TITLE   = Font(name="Arial", size=14, bold=True)
FONT_SECTION = Font(name="Arial", size=12, bold=True)

THIN = Side(border_style="thin", color="BDBDBD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    "decisión",
    "proba_v2", "cosine", "cluster",
    "easy_marca", "easy_título", "easy_precio",
    "sodi_marca", "sodi_título", "sodi_precio",
    "easy_sku", "sodi_sku", "origen",
]
WIDTHS = [11, 9, 8, 9, 14, 55, 12, 14, 55, 12, 14, 14, 14]


def _header(ws, row, headers):
    fill = PatternFill("solid", fgColor=COLOR_HEADER)
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 26


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fila(rec, origen=""):
    return [
        "",
        rec.get("proba", rec.get("proba_v2")),
        rec["cosine"], rec["cluster_id"],
        rec.get("easy_marca") or "",
        rec.get("easy_titulo") or "",
        rec.get("easy_precio"),
        rec.get("sodimac_marca") or "",
        rec.get("sodimac_titulo") or "",
        rec.get("sodimac_precio"),
        rec.get("easy_sku") or "",
        rec.get("sodimac_sku") or "",
        origen,
    ]


def _escribir(ws, rows, start=2, color=None, origen=""):
    for i, r in enumerate(rows, start=start):
        vals = _fila(r, origen=origen or r.get("_origen", ""))
        fill = PatternFill("solid", fgColor=color) if color else None
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_DEFAULT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            if fill:
                c.fill = fill
    n = len(rows)
    if n:
        dv = DataValidation(type="list", formula1='"✓ match,✗ no match,? duda"', allow_blank=True)
        dv.add(f"A{start}:A{start+n-1}")
        ws.add_data_validation(dv)


def _hoja_resumen(wb, compare, v2_top, v2_cands):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    r = 1
    ws.cell(row=r, column=1, value="MATCHER V2 - PANEL DE REVISIÓN").font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 2

    ws.cell(row=r, column=1, value="COMPARACIÓN V1 vs V2").font = FONT_SECTION
    r += 1
    _header(ws, r, ["métrica", "V1 (leaky)", "V2 (v1+v2)"])
    r += 1
    filas = [
        ("Matches 1-a-1 alta confianza", compare["v1"]["matches_1a1"], compare["v2"]["matches_1a1"]),
        ("Ambiguos (ties)",              compare["v1"]["ambiguos"],    compare["v2"]["ambiguos"]),
        ("Candidatos totales",           compare["v1"]["candidatos"],  compare["v2"]["candidatos"]),
    ]
    for k, a, b in filas:
        ws.cell(row=r, column=1, value=k).font = FONT_DEFAULT
        ws.cell(row=r, column=2, value=a).font = FONT_DEFAULT
        ws.cell(row=r, column=3, value=b).font = FONT_DEFAULT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="OVERLAP V1 ∩ V2 (1-a-1)").font = FONT_SECTION
    r += 1
    ov = compare["overlap"]
    ws.cell(row=r, column=1, value="En ambos modelos").font = FONT_DEFAULT
    ws.cell(row=r, column=2, value=ov["en_ambos"]).font = FONT_DEFAULT
    r += 1
    ws.cell(row=r, column=1, value="Solo V1 (V2 rechazó)").font = FONT_DEFAULT
    ws.cell(row=r, column=2, value=ov["solo_v1"]).font = FONT_DEFAULT
    r += 1
    ws.cell(row=r, column=1, value="Solo V2 (nuevos matches)").font = FONT_DEFAULT
    ws.cell(row=r, column=2, value=ov["solo_v2"]).font = FONT_DEFAULT
    r += 2

    ws.cell(row=r, column=1, value="VALIDACIÓN ORTOGONAL DEL MODELO V2").font = FONT_SECTION
    r += 1
    mensaje = (
        "El modelo V2 fue validado con test ortogonal:\n"
        "  • Train=WEAK → Test=GOLDEN: F1 = 0.947, AUC = 0.995\n"
        "  • Train=GOLDEN → Test=WEAK: F1 = 0.551, AUC = 0.875 (precision = 1.000)\n\n"
        "Feature importance: v1=58%, v2=42%  (cosine_sim bajó de 52% a 29%).\n"
        "Conclusión: el modelo V2 generaliza más allá de memorizar las reglas.\n"
        "Aún así requiere validación humana: persisten falsos positivos donde marca\n"
        "coincide pero el producto difiere (ej. Dewalt cargador vs Dewalt atornillador)."
    )
    c = ws.cell(row=r, column=1, value=mensaje)
    c.font = FONT_DEFAULT
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r+6, end_column=3)
    r += 8

    ws.cell(row=r, column=1, value="HOJAS").font = FONT_SECTION
    r += 1
    hojas = [
        ("Matches_V2_1a1", "34 matches de alta confianza del modelo V2 (verde)"),
        ("Solo_V2_nuevos", "31 matches que V2 encontró y V1 no (ámbar — alto interés)"),
        ("Solo_V1_rechazados", "5 matches V1 que V2 rechazó (rojo — validar con cuidado)"),
        ("Ambiguos_V2", "Primeros 300 pares con ties (naranja — decisión humana)"),
        ("Top_candidatos_V2", "500 mejores candidatos V2 (por proba descendente)"),
    ]
    for n, d in hojas:
        ws.cell(row=r, column=1, value=n).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=2, value=d).font = FONT_DEFAULT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1


def main():
    print("=" * 70)
    print("EXPORTANDO MATCHES V2 A EXCEL")
    print("=" * 70)

    with open(IN_V2_CANDS, encoding="utf-8") as f:
        v2_cands = json.load(f)
    with open(IN_V2_TOP, encoding="utf-8") as f:
        v2_top = json.load(f)
    with open(IN_V1_TOP, encoding="utf-8") as f:
        v1_top = json.load(f)
    with open(IN_COMPARE, encoding="utf-8") as f:
        compare = json.load(f)

    # sets de pares V1 y V2
    v1_pairs = {(m["easy_id"], m["sodimac_id"]): m for m in v1_top["matches_1a1"]}
    v2_pairs = {(m["easy_id"], m["sodimac_id"]): m for m in v2_top["matches_1a1"]}

    solo_v2 = [m for k, m in v2_pairs.items() if k not in v1_pairs]
    solo_v1 = [m for k, m in v1_pairs.items() if k not in v2_pairs]
    en_ambos = [m for k, m in v2_pairs.items() if k in v1_pairs]

    # todos ordenados por proba
    for m in v2_top["matches_1a1"]:
        m["_origen"] = "V2∩V1" if (m["easy_id"], m["sodimac_id"]) in v1_pairs else "solo_V2"
    for m in solo_v1:
        m["_origen"] = "solo_V1"

    v2_top["matches_1a1"].sort(key=lambda r: -r["proba"])
    solo_v2.sort(key=lambda r: -r["proba"])

    wb = Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, compare, v2_top, v2_cands)

    # matches V2 1-a-1 completos (34)
    ws = wb.create_sheet("Matches_V2_1a1")
    _widths(ws, WIDTHS); _header(ws, 1, COLS); ws.freeze_panes = "A2"
    _escribir(ws, v2_top["matches_1a1"], color=COLOR_MATCH)

    # solo V2 (los nuevos — alto interés)
    ws = wb.create_sheet("Solo_V2_nuevos")
    _widths(ws, WIDTHS); _header(ws, 1, COLS); ws.freeze_panes = "A2"
    _escribir(ws, solo_v2, color=COLOR_NEW_V2)

    # solo V1 (los que V2 descartó)
    ws = wb.create_sheet("Solo_V1_rechazados")
    _widths(ws, WIDTHS); _header(ws, 1, COLS); ws.freeze_panes = "A2"
    _escribir(ws, solo_v1, color=COLOR_SOLO_V1, origen="solo_V1")

    # ambiguos V2 (top 300)
    amb = v2_top.get("ambiguos", [])
    amb.sort(key=lambda r: -r["proba"])
    ws = wb.create_sheet("Ambiguos_V2")
    _widths(ws, WIDTHS); _header(ws, 1, COLS); ws.freeze_panes = "A2"
    _escribir(ws, amb[:300], color=COLOR_AMB)

    # top 500 candidatos
    v2_cands.sort(key=lambda r: -r["proba"])
    ws = wb.create_sheet("Top_candidatos_V2")
    _widths(ws, WIDTHS); _header(ws, 1, COLS); ws.freeze_panes = "A2"
    # colorear por proba
    for i, r in enumerate(v2_cands[:500], start=2):
        color = COLOR_ALTA if r["proba"] >= 0.90 else COLOR_MEDIA
        fill = PatternFill("solid", fgColor=color)
        vals = _fila(r)
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=v)
            c.font = FONT_DEFAULT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            c.fill = fill
    dv = DataValidation(type="list", formula1='"✓ match,✗ no match,? duda"', allow_blank=True)
    dv.add(f"A2:A{min(501, len(v2_cands)+1)}")
    ws.add_data_validation(dv)

    wb.save(OUT_XLSX)
    size_kb = OUT_XLSX.stat().st_size // 1024
    print(f"\n  Excel : {OUT_XLSX}")
    print(f"  Tamaño: {size_kb} KB")
    print(f"  Hojas : {wb.sheetnames}")


if __name__ == "__main__":
    main()
