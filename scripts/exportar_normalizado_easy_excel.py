"""Exporta catalogo_normalizado.json de Easy a Excel."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA   = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data\easy")
SALIDA = DATA / "catalogo_normalizado_easy.xlsx"

# ---------------------------------------------------------------------------
# Paleta Easy (verde-azul)
# ---------------------------------------------------------------------------
C_DARK  = "154360"
C_MID   = "1F618D"
C_LIGHT = "D6EAF8"
C_WHITE = "FFFFFF"
C_BORD  = "A9CCE3"
C_GREEN = "196F3D"   # para largo/diametro (datos técnicos clave)
C_GRLT  = "D5F5E3"

FONT_H    = Font(name="Arial", bold=True, color=C_WHITE, size=10)
FONT_B    = Font(name="Arial", size=9)
FONT_BOLD = Font(name="Arial", size=9, bold=True)
FONT_LINK = Font(name="Arial", size=9, color="0563C1", underline="single")
FONT_NUM  = Font(name="Arial", size=9, color="1A5276")

FILL_DARK  = PatternFill("solid", fgColor=C_DARK)
FILL_MID   = PatternFill("solid", fgColor=C_MID)
FILL_LIGHT = PatternFill("solid", fgColor=C_LIGHT)
FILL_GREEN = PatternFill("solid", fgColor=C_GREEN)
FILL_GRLT  = PatternFill("solid", fgColor=C_GRLT)

AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left",   vertical="center")
AL_W = Alignment(horizontal="left",   vertical="top", wrap_text=True)

thin   = Side(style="thin", color=C_BORD)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr(ws, row, col, texto, fill=None, span=None):
    cell = ws.cell(row=row, column=col)
    cell.value = texto
    cell.font  = FONT_H
    cell.fill  = fill or FILL_DARK
    cell.alignment = AL_C
    cell.border = BORDER
    if span and span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)


def cel(ws, row, col, valor, fmt=None, link=False, bold=False,
        wrap=False, zfill=None, center=False):
    cell = ws.cell(row=row, column=col)
    cell.value = valor
    if link:
        cell.font = FONT_LINK
    elif bold:
        cell.font = FONT_BOLD
    elif fmt:
        cell.font = FONT_NUM
    else:
        cell.font = FONT_B
    cell.alignment = AL_W if wrap else (AL_C if center else AL_L)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if zfill:
        cell.fill = zfill


# ---------------------------------------------------------------------------
def main():
    with open(DATA / "catalogo_normalizado.json", encoding="utf-8") as f:
        data = json.load(f)

    n = len(data)
    wb = Workbook()

    # =======================================================================
    # HOJA 1 — Catálogo normalizado
    # =======================================================================
    ws = wb.active
    ws.title = "Catálogo normalizado"
    ws.freeze_panes = "A3"

    # --- Fila 1: grupos ---
    # Col:  1    2   3   4   5   6   7 | 8  9  10  11  12  13 | 14  15  16  17  18  19  20  21 | 22  23 | 24  25
    grupos = [
        ("ID & BÁSICO",          1,  7,  FILL_DARK),
        ("PRECIOS",              8, 10,  FILL_MID),
        ("CATEGORÍAS",          11, 14,  FILL_DARK),
        ("DIMENSIONES",         15, 17,  FILL_GREEN),
        ("ATRIBUTOS TÉCNICOS",  18, 23,  FILL_MID),
        ("CONTENIDO",           24, 25,  FILL_DARK),
    ]
    for etiqueta, c1, c2, fill in grupos:
        hdr(ws, 1, c1, etiqueta, fill, span=c2 - c1 + 1)

    # --- Fila 2: encabezados columna ---
    headers = [
        # ID & BÁSICO (1-7)
        "ID Producto", "Tienda", "SKU", "Marca", "Título",
        "URL Producto", "URL Imagen",
        # PRECIOS (8-10)
        "Precio CLP", "Precio Normal CLP", "Descuento %",
        # CATEGORÍAS (11-14)
        "Categoría 1", "Categoría 2", "Categoría 3", "Categoría 4",
        # DIMENSIONES (15-17)
        "Largo (mm)", "Diámetro (mm)", "Medida cruda",
        # ATRIBUTOS TÉCNICOS (18-23)
        "Material", "Tipo cabeza", "Tipo rosca", "Tipo punta",
        "Color", "Cantidad empaque",
        # CONTENIDO (24-25)
        "EAN", "Disponibilidad",
    ]

    fills_h2 = (
        [FILL_DARK] * 7 + [FILL_MID] * 3 + [FILL_DARK] * 4
        + [FILL_GREEN] * 3 + [FILL_MID] * 6 + [FILL_DARK] * 2
    )
    for col, (h, fill) in enumerate(zip(headers, fills_h2), 1):
        hdr(ws, 2, col, h, fill)

    # --- Filas de datos ---
    for fi, prod in enumerate(data, 3):
        zf = FILL_LIGHT if fi % 2 == 0 else None
        mb = prod.get("metadata_basica") or {}
        mt = prod.get("metadata_tecnica") or {}
        dims = mt.get("dimensiones") or {}
        cats = mb.get("categorias") or []

        precio_actual = mb.get("precio_clp") or 0
        precio_normal = mb.get("precio_normal_clp") or precio_actual
        desc_pct      = mb.get("descuento_pct")

        # ID & BÁSICO
        cel(ws, fi, 1,  prod.get("id_producto", ""),   bold=True, zfill=zf)
        cel(ws, fi, 2,  prod.get("tienda", ""),         center=True, zfill=zf)
        cel(ws, fi, 3,  str(prod.get("sku", "")),       bold=True, zfill=zf)
        cel(ws, fi, 4,  mb.get("marca", ""),            zfill=zf)
        cel(ws, fi, 5,  mb.get("titulo", ""),           zfill=zf)
        cel(ws, fi, 6,  prod.get("url_producto", ""),   link=True,  zfill=zf)
        cel(ws, fi, 7,  prod.get("url_imagen", ""),     link=True,  zfill=zf)

        # PRECIOS
        c_precio = ws.cell(row=fi, column=8)
        c_precio.value = int(precio_actual) if precio_actual else 0
        c_precio.number_format = "#,##0"
        c_precio.font = FONT_NUM
        c_precio.alignment = AL_C
        c_precio.border = BORDER
        if zf: c_precio.fill = zf

        c_normal = ws.cell(row=fi, column=9)
        c_normal.value = int(precio_normal) if precio_normal else 0
        c_normal.number_format = "#,##0"
        c_normal.font = FONT_NUM
        c_normal.alignment = AL_C
        c_normal.border = BORDER
        if zf: c_normal.fill = zf

        c_desc = ws.cell(row=fi, column=10)
        c_desc.value = desc_pct if desc_pct else ""
        c_desc.number_format = '0.0"%"'
        c_desc.font = FONT_NUM
        c_desc.alignment = AL_C
        c_desc.border = BORDER
        if zf: c_desc.fill = zf

        # CATEGORÍAS
        for ci, cat in enumerate(cats[:4], 11):
            cel(ws, fi, ci, cat, zfill=zf)
        for ci in range(len(cats[:4]) + 11, 15):
            cel(ws, fi, ci, "", zfill=zf)

        # DIMENSIONES — zebra con tono verde claro
        zf_dim = FILL_GRLT if fi % 2 == 0 else None
        largo = dims.get("largo_mm")
        diam  = dims.get("diametro_mm")

        c_l = ws.cell(row=fi, column=15)
        c_l.value = largo if largo else ""
        c_l.number_format = "0.00"
        c_l.font = FONT_NUM
        c_l.alignment = AL_C
        c_l.border = BORDER
        if zf_dim: c_l.fill = zf_dim

        c_d = ws.cell(row=fi, column=16)
        c_d.value = diam if diam else ""
        c_d.number_format = "0.00"
        c_d.font = FONT_NUM
        c_d.alignment = AL_C
        c_d.border = BORDER
        if zf_dim: c_d.fill = zf_dim

        cel(ws, fi, 17, dims.get("medida_cruda", ""), zfill=zf)

        # ATRIBUTOS TÉCNICOS
        cel(ws, fi, 18, mt.get("material", ""),       zfill=zf)
        cel(ws, fi, 19, mt.get("tipo_cabeza", ""),    zfill=zf)
        cel(ws, fi, 20, mt.get("tipo_rosca", ""),     zfill=zf)
        cel(ws, fi, 21, mt.get("tipo_punta", ""),     zfill=zf)
        cel(ws, fi, 22, mt.get("color", ""),          zfill=zf)

        c_cant = ws.cell(row=fi, column=23)
        cant = mt.get("cantidad_empaque") or 1
        c_cant.value = int(cant) if cant else 1
        c_cant.number_format = "#,##0"
        c_cant.font = FONT_NUM
        c_cant.alignment = AL_C
        c_cant.border = BORDER
        if zf: c_cant.fill = zf

        # CONTENIDO
        cel(ws, fi, 24, mb.get("ean", ""),            zfill=zf)
        cel(ws, fi, 25, mb.get("disponibilidad", ""), center=True, zfill=zf)

    # Anchos
    anchos = {
        1: 24, 2: 9, 3: 13, 4: 16, 5: 52,
        6: 42, 7: 42,
        8: 15, 9: 16, 10: 13,
        11: 26, 12: 26, 13: 26, 14: 26,
        15: 12, 16: 14, 17: 18,
        18: 14, 19: 14, 20: 14, 21: 14, 22: 14, 23: 16,
        24: 16, 25: 14,
    }
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 34

    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

    # =======================================================================
    # HOJA 2 — Resumen
    # =======================================================================
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 20

    hdr(ws2, 1, 1, "Métrica", FILL_DARK)
    hdr(ws2, 1, 2, "Valor",   FILL_DARK)
    ws2.row_dimensions[1].height = 24

    metricas = [
        ("Total productos",                   f"=COUNTA('Catálogo normalizado'!C3:C{n+2})"),
        ("Marcas únicas",                     f"=SUMPRODUCT(1/COUNTIF('Catálogo normalizado'!D3:D{n+2},'Catálogo normalizado'!D3:D{n+2}))"),
        ("Precio promedio (CLP)",             f"=AVERAGE('Catálogo normalizado'!H3:H{n+2})"),
        ("Precio mínimo (CLP)",               f"=MIN('Catálogo normalizado'!H3:H{n+2})"),
        ("Precio máximo (CLP)",               f"=MAX('Catálogo normalizado'!H3:H{n+2})"),
        ("Productos InStock",                 f"=COUNTIF('Catálogo normalizado'!Y3:Y{n+2},\"InStock\")"),
        ("Productos OutOfStock",              f"=COUNTIF('Catálogo normalizado'!Y3:Y{n+2},\"OutOfStock\")"),
        ("Productos con EAN",                 f"=COUNTA('Catálogo normalizado'!X3:X{n+2})"),
        ("Productos con largo_mm extraído",   f"=COUNTA('Catálogo normalizado'!O3:O{n+2})"),
        ("Productos con diámetro_mm extraído",f"=COUNTA('Catálogo normalizado'!P3:P{n+2})"),
        ("Productos con material",            f"=COUNTA('Catálogo normalizado'!R3:R{n+2})"),
        ("Productos con categorías",          f"=COUNTA('Catálogo normalizado'!K3:K{n+2})"),
        ("Cantidad promedio por empaque",     f"=AVERAGE('Catálogo normalizado'!W3:W{n+2})"),
    ]

    for fi, (etiqueta, valor) in enumerate(metricas, 2):
        ca = ws2.cell(row=fi, column=1)
        cb = ws2.cell(row=fi, column=2)
        ca.value = etiqueta
        ca.font  = FONT_B
        ca.border = BORDER
        ca.alignment = AL_L
        cb.value = valor
        cb.font  = FONT_NUM
        cb.border = BORDER
        cb.alignment = AL_L
        if "precio" in etiqueta.lower() or "promedio" in etiqueta.lower():
            cb.number_format = "#,##0"
        if fi % 2 == 0:
            ca.fill = FILL_LIGHT
            cb.fill = FILL_LIGHT

    # =======================================================================
    wb.save(SALIDA)
    print(f"Excel guardado: {SALIDA}")
    print(f"Productos: {n} | Columnas: {len(headers)}")


if __name__ == "__main__":
    main()
