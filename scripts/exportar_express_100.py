"""Genera el Excel 'Express 100' para revisión humana rápida (~30 min).

Composición estratégica:
  - Hoja única con ~100 filas ordenadas por prioridad
  - Sección A: Matches_V2_1a1       (34) → alta confianza, muchos TP reales
  - Sección B: Solo_V1_rechazados   ( 5) → ¿acertó V2 al descartarlos?
  - Sección C: Solo_V2_nuevos       (31) → descubrimientos nuevos del modelo
  - Sección D: Ambiguos_V2 top 30   (30) → borderline más valiosos (proba>=0.85)

Total: exactamente 100 filas (si hay suficientes ambiguos con proba>=0.85)
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE       = Path(r"C:\Users\Administrator\Documents\Buscop\motordelco-o\data")
IN_V2_TOP  = BASE / "matches_v2_top.json"
IN_V1_TOP  = BASE / "matches_top.json"
OUT_XLSX   = BASE / "express_100_revisar.xlsx"

# ── Colores por sección ────────────────────────────────────────────────────
COLOR_HDR_A  = "1B5E20"   # verde oscuro – sección A (matches 1-a-1)
COLOR_HDR_B  = "B71C1C"   # rojo oscuro  – sección B (solo V1 rechazados)
COLOR_HDR_C  = "E65100"   # naranja osc. – sección C (solo V2 nuevos)
COLOR_HDR_D  = "1A237E"   # azul oscuro  – sección D (ambiguos top)

COLOR_ROW_A  = "E8F5E9"   # verde muy suave
COLOR_ROW_B  = "FFEBEE"   # rojo muy suave
COLOR_ROW_C  = "FFF3E0"   # naranja muy suave
COLOR_ROW_D  = "E8EAF6"   # azul muy suave

COLOR_GLOBAL_HDR = "37474F"

THIN  = Side(border_style="thin", color="BDBDBD")
THICK = Side(border_style="medium", color="455A64")
BORDER_THIN   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOP    = Border(left=THIN, right=THIN, top=THICK, bottom=THIN)

FONT_DEFAULT  = Font(name="Arial", size=10)
FONT_HEADER   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_SECTION  = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_TITLE    = Font(name="Arial", size=14, bold=True)
FONT_INSTR    = Font(name="Arial", size=10, italic=True, color="455A64")
FONT_NUM      = Font(name="Arial", size=10, bold=True)

COLS = [
    "#", "decisión",
    "proba", "cosine", "cluster",
    "easy_marca",  "easy_título",   "easy_precio",
    "sodi_marca",  "sodi_título",   "sodi_precio",
    "easy_sku",    "sodi_sku",
]
WIDTHS = [5, 12, 8, 8, 9, 14, 55, 12, 14, 55, 12, 14, 14]

SECCIONES = {
    "A": {
        "hdr_color": COLOR_HDR_A, "row_color": COLOR_ROW_A,
        "titulo": "SECCIÓN A — Matches 1-a-1 alta confianza (V2)",
        "desc":   "El modelo tiene alta confianza en estos pares. Se espera mayoría de TP.",
    },
    "B": {
        "hdr_color": COLOR_HDR_B, "row_color": COLOR_ROW_B,
        "titulo": "SECCIÓN B — V1 los aceptó, V2 los rechazó",
        "desc":   "Solo 5 pares. Valida si V2 acertó al ser más conservador.",
    },
    "C": {
        "hdr_color": COLOR_HDR_C, "row_color": COLOR_ROW_C,
        "titulo": "SECCIÓN C — Solo en V2 (descubrimientos nuevos)",
        "desc":   "Matches que V1 no encontró. Alto interés, pero posible FP.",
    },
    "D": {
        "hdr_color": COLOR_HDR_D, "row_color": COLOR_ROW_D,
        "titulo": "SECCIÓN D — Ambiguos borderline (proba >= 0.85)",
        "desc":   "El modelo dudó entre 2 candidatos. Tu decisión rompe el empate.",
    },
}


def _set_widths(ws):
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _fila_datos(n, rec):
    return [
        n,
        "",
        rec.get("proba"),
        rec.get("cosine"),
        rec.get("cluster_id"),
        rec.get("easy_marca") or "",
        rec.get("easy_titulo") or "",
        rec.get("easy_precio"),
        rec.get("sodimac_marca") or "",
        rec.get("sodimac_titulo") or "",
        rec.get("sodimac_precio"),
        rec.get("easy_sku") or "",
        rec.get("sodimac_sku") or "",
    ]


def _escribir_col_header(ws, row):
    fill = PatternFill("solid", fgColor=COLOR_GLOBAL_HDR)
    for col, h in enumerate(COLS, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[row].height = 24


def _escribir_seccion_header(ws, row, sec_id, n_filas):
    cfg = SECCIONES[sec_id]
    fill = PatternFill("solid", fgColor=cfg["hdr_color"])
    # fila de título de sección
    c = ws.cell(row=row, column=1,
                value=f"  {cfg['titulo']}  ({n_filas} filas)")
    c.font = FONT_SECTION
    c.fill = fill
    c.alignment = Alignment(vertical="center")
    c.border = BORDER_TOP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    ws.row_dimensions[row].height = 22
    # fila de descripción
    row += 1
    c = ws.cell(row=row, column=1, value=f"  ℹ  {cfg['desc']}")
    c.font = FONT_INSTR
    c.alignment = Alignment(vertical="center")
    c.border = BORDER_THIN
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    ws.row_dimensions[row].height = 18
    return row + 1


def _escribir_filas(ws, rows_data, sec_id, start_row, contador_inicio):
    cfg = SECCIONES[sec_id]
    fill_row = PatternFill("solid", fgColor=cfg["row_color"])

    row = start_row
    for i, rec in enumerate(rows_data):
        vals = _fila_datos(contador_inicio + i, rec)
        for col, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = FONT_NUM if col == 1 else FONT_DEFAULT
            c.alignment = Alignment(vertical="center", wrap_text=(col in (7, 10)),
                                    horizontal="center" if col in (1, 3, 4, 5, 8, 11) else "left")
            c.border = BORDER_THIN
            c.fill = fill_row
        ws.row_dimensions[row].height = 30
        row += 1
    return row


def main():
    print("=" * 70)
    print("GENERANDO EXPRESS_100_REVISAR.XLSX")
    print("=" * 70)

    with open(IN_V2_TOP, encoding="utf-8") as f:
        v2_top = json.load(f)
    with open(IN_V1_TOP, encoding="utf-8") as f:
        v1_top = json.load(f)

    # construir sets de pares
    v1_pairs = {(m["easy_id"], m["sodimac_id"]) for m in v1_top["matches_1a1"]}
    v2_pairs = {(m["easy_id"], m["sodimac_id"]) for m in v2_top["matches_1a1"]}

    sec_a = sorted(v2_top["matches_1a1"], key=lambda r: -r["proba"])
    sec_b = [m for m in v1_top["matches_1a1"] if (m["easy_id"], m["sodimac_id"]) not in v2_pairs]
    sec_c = sorted(
        [m for m in v2_top["matches_1a1"] if (m["easy_id"], m["sodimac_id"]) not in v1_pairs],
        key=lambda r: -r["proba"],
    )
    # ambiguos: tomar los 30 con mayor proba (ya vienen del modelo v2)
    ambiguos = sorted(v2_top.get("ambiguos", []), key=lambda r: -r["proba"])
    # filtrar: solo proba >= 0.85, máximo 30
    sec_d = [a for a in ambiguos if a.get("proba", 0) >= 0.85][:30]
    # si no hay suficientes con 0.85, bajar el umbral
    if len(sec_d) < 10:
        sec_d = ambiguos[:30]

    total = len(sec_a) + len(sec_b) + len(sec_c) + len(sec_d)
    print(f"  Sección A (matches 1-a-1 V2) : {len(sec_a):>3}")
    print(f"  Sección B (solo V1 rechazados): {len(sec_b):>3}")
    print(f"  Sección C (solo V2 nuevos)    : {len(sec_c):>3}")
    print(f"  Sección D (ambiguos top)      : {len(sec_d):>3}")
    print(f"  TOTAL                         : {total:>3}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Express_100"
    _set_widths(ws)
    ws.freeze_panes = "A5"  # congela título + subtítulo + encabezados col

    # ── Cabecera del documento ──────────────────────────────────────────────
    r = 1
    c = ws.cell(row=r, column=1,
                value="EXPRESS 100 — Revisión de Matches Easy × Sodimac")
    c.font = FONT_TITLE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    ws.row_dimensions[r].height = 28

    r = 2
    instrucciones = (
        "INSTRUCCIONES: En la columna 'decisión' elige  ✓ match  |  ✗ no match  |  ? duda   "
        "para cada par. Ordena por sección (A → D). "
        "Si el precio y las dimensiones coinciden = señal fuerte de match."
    )
    c = ws.cell(row=r, column=1, value=instrucciones)
    c.font = FONT_INSTR
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    ws.row_dimensions[r].height = 32

    # ── Encabezados de columna ─────────────────────────────────────────────
    r = 3
    _escribir_col_header(ws, r)
    r = 4

    # ── Secciones ──────────────────────────────────────────────────────────
    contador = 1
    dv_ranges = []

    for sec_id, datos in [("A", sec_a), ("B", sec_b), ("C", sec_c), ("D", sec_d)]:
        if not datos:
            continue
        r = _escribir_seccion_header(ws, r, sec_id, len(datos))
        data_start = r
        r = _escribir_filas(ws, datos, sec_id, r, contador)
        dv_ranges.append(f"B{data_start}:B{r-1}")
        contador += len(datos)

    # ── Data validation en toda la columna B (decisión) ───────────────────
    dv = DataValidation(
        type="list",
        formula1='"✓ match,✗ no match,? duda"',
        allow_blank=True,
        showDropDown=False,
    )
    for rng in dv_ranges:
        dv.add(rng)
    ws.add_data_validation(dv)

    wb.save(OUT_XLSX)
    size_kb = OUT_XLSX.stat().st_size // 1024
    print(f"\n  Guardado : {OUT_XLSX}")
    print(f"  Tamaño   : {size_kb} KB")
    print(f"  Total filas de datos: {total}")
    print(f"\n  Tiempo estimado de revisión: ~{total // 4}–{total // 3} minutos")


if __name__ == "__main__":
    main()
